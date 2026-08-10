#!/usr/bin/env python3
"""
phishing_toolkit.py — Bộ công cụ hỗ trợ phát hiện & báo cáo domain phishing.

3 lệnh con:

  check <domain>
      Kiểm tra đầy đủ 1 domain: SSL issuer/serial, WHOIS, Cloudflare,
      VirusTotal, Google Safe Browsing (nếu có API key), ghi log vào
      case_log.csv, và sinh sẵn email báo cáo trong thư mục reports/.

  related <keyword>
      Tìm các domain khác có chứng chỉ SSL chứa từ khóa (vd tên thương hiệu)
      qua Certificate Transparency log (crt.sh) — giúp phát hiện các domain
      "anh em" cùng chiến dịch phishing.

  brandscan <domain>
      Sinh các biến thể gõ nhầm/giống domain thật (dùng dnstwist) và lọc ra
      những cái đã bị đăng ký — để chủ động phát hiện phishing trước khi có
      người báo cáo.

Cài đặt:
    pip install requests python-whois cryptography dnstwist

Cấu hình:
    Copy config.example.ini -> config.ini và điền VT/GSB API key (tùy chọn).
"""

import argparse
import concurrent.futures
import configparser
import csv
import imaplib
import json
import os
import random
import re
import smtplib
import socket
import ssl
import subprocess
import sys
import time
from contextlib import contextmanager
from datetime import datetime, timezone
from email.mime.text import MIMEText
from urllib.parse import urlparse

try:
    import whois
except ImportError:
    whois = None

try:
    import requests
    requests.packages.urllib3.disable_warnings()  # detect_cdn() gọi HEAD với verify=False cho domain nghi vấn
except ImportError:
    requests = None

try:
    import dns.resolver
except ImportError:
    dns = None

try:
    # IP WHOIS (registry ARIN/RIPE/APNIC... qua RDAP) — KHÁC hẳn `whois` ở trên (domain WHOIS).
    # Dùng ipwhois thay vì shell ra lệnh `whois <IP>` như 03_Technical_Guide.md gợi ý vì lệnh
    # `whois` không có sẵn mặc định trên Windows — sẽ lặp lại đúng vấn đề PATH đã gặp với
    # streamlit/dnstwist. ipwhois dùng RDAP thuần Python qua mạng, không cần binary ngoài.
    from ipwhois import IPWhois
except ImportError:
    IPWhois = None

try:
    # PySocks — chỉ cần khi cấu hình proxy cho SMTP (smtp_proxies trong config.ini).
    # Nếu không cài, gửi email vẫn chạy bình thường miễn là smtp_proxies để trống.
    import socks as _socks
    _SOCKS5 = _socks.SOCKS5
    _SOCKS4 = _socks.SOCKS4
    _SOCKS_HTTP = _socks.HTTP
except ImportError:
    _socks = None
    _SOCKS5 = _SOCKS4 = _SOCKS_HTTP = None


# Khi chạy từ .exe (PyInstaller frozen), __file__ trỏ vào thư mục temp _MEIPASS
# nên config.ini / reports / logs phải đặt cạnh file .exe, không phải trong _MEIPASS.
if getattr(sys, "frozen", False):
    # Thử sys.executable trước (thư mục chứa .exe), fallback cwd
    _exe_dir = os.path.dirname(sys.executable)
    _cwd = os.getcwd()
    # Ưu tiên thư mục nào có config.ini
    if os.path.exists(os.path.join(_exe_dir, "config.ini")):
        BASE_DIR = _exe_dir
    elif os.path.exists(os.path.join(_cwd, "config.ini")):
        BASE_DIR = _cwd
    else:
        BASE_DIR = _exe_dir  # default: cạnh .exe, config.ini sẽ được tạo ở đây
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.ini")
LOG_PATH = os.path.join(BASE_DIR, "case_log.csv")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
SENT_LOG_PATH = os.path.join(BASE_DIR, "sent_log.csv")

CA_ABUSE_NOTES = {
    "google trust services": {
        "report_url": None,
        "abuse_email": None,
        "note": (
            "Google Trust Services KHÔNG xử lý report phishing/malware — chỉ xử lý mis-issuance. "
            "Bỏ qua bước này."
        ),
        "can_revoke": False,
    },
    "let's encrypt": {
        "report_url": "https://letsencrypt.org/abuse/",
        "abuse_email": "abuse@letsencrypt.org",
        "note": "Let's Encrypt xử lý report qua form hoặc email. Xác suất revoke cert thấp hơn các CA thương mại.",
        "can_revoke": True,
    },
    "sectigo": {
        "report_url": "https://sectigo.com/support-resources/report-abuse-ssl-certificate",
        "abuse_email": None,
        "note": "Sectigo có xử lý report phishing/malware — dùng web form.",
        "can_revoke": True,
    },
    "comodo": {
        "report_url": "https://sectigo.com/support-resources/report-abuse-ssl-certificate",
        "abuse_email": None,
        "note": "Comodo CA (nay là Sectigo) — dùng form Sectigo.",
        "can_revoke": True,
    },
    "digicert": {
        "report_url": "https://www.digicert.com/reporting-abuse/",
        "abuse_email": "abuse@digicert.com",
        "note": "DigiCert có xử lý report phishing/malware — form hoặc email.",
        "can_revoke": True,
    },
    "zerossl": {
        "report_url": None,
        "abuse_email": "abuse@zerossl.com",
        "note": "ZeroSSL nhận report qua email.",
        "can_revoke": True,
    },
    "godaddy": {
        "report_url": "https://sg.godaddy.com/help/report-abuse-24108",
        "abuse_email": "abuse@godaddy.com",
        "note": "GoDaddy (vai trò CA) có xử lý report phishing — form hoặc email.",
        "can_revoke": True,
    },
    "globalsign": {
        "report_url": "https://www.globalsign.com/en/company/legal/abuse/",
        "abuse_email": "abuse@globalsign.com",
        "note": "GlobalSign có xử lý report phishing/malware.",
        "can_revoke": True,
    },
    "entrust": {
        "report_url": None,
        "abuse_email": "certissues@entrust.com",
        "note": "Entrust nhận report qua email.",
        "can_revoke": True,
    },
    "actalis": {
        "report_url": None,
        "abuse_email": "tls-abuse@actalis.it",
        "note": "Actalis nhận report qua email.",
        "can_revoke": True,
    },
}

# CDN/anti-DDoS proxy phổ biến hay bị phishing lợi dụng để ẩn IP gốc (03_Technical_Guide.md mục 2).
# "cloudflare" ở đây chỉ cung cấp report_url/note dùng chung — việc PHÁT HIỆN có dùng Cloudflare hay
# không vẫn do is_cloudflare() (tra nameserver) đảm nhiệm, không phải detect_cdn().
CDN_ABUSE_CONTACTS = {
    "cloudflare": {
        "report_url": "https://abuse.cloudflare.com",
        "note": "Chọn mục \"Phishing & Malware\" khi report — có thể chặn proxy hoặc gắn interstitial cảnh báo.",
    },
    "fastly": {
        "report_url": "https://www.fastly.com/abuse",
        "note": "Report qua form web.",
    },
    "akamai": {
        "report_url": "https://www.akamai.com/legal/compliance/report-abuse",
        "note": "Report qua form compliance.",
    },
    "cloudfront": {
        "report_url": "https://support.aws.amazon.com/#/contacts/report-abuse",
        "note": "CDN của AWS — dùng form Trust & Safety của AWS.",
    },
    "stormwall": {
        "report_url": "abuse@stormwall.pro",
        "note": "Cũng có thể report qua live chat trên trang chủ Stormwall.",
    },
    "ddos-guard": {
        "report_url": "https://ddos-guard.net/en/abuse",
        "note": "Anti-DDoS proxy — thường ẩn IP gốc, kết hợp quét subdomain (origin_ip_scan) để tìm IP thật.",
    },
    # Hosting platform / cloud — phishing thường host trực tiếp trên đây
    "vercel": {
        "report_url": "https://vercel.com/abuse",
        "note": "Vercel hosting — điền web form.",
    },
    "netlify": {
        "report_url": "fraud@netlify.com",
        "note": "Netlify hosting — chỉ nhận qua email fraud@netlify.com.",
    },
    "github-pages": {
        "report_url": "https://support.github.com/contact/report-abuse",
        "note": "GitHub Pages — report qua form support.",
    },
    "digitalocean": {
        "report_url": "https://www.digitalocean.com/company/contact/abuse",
        "note": "DigitalOcean VPS/Droplet.",
    },
    "hetzner": {
        "report_url": "https://abuse.hetzner.com",
        "note": "Hetzner Cloud/dedicated.",
    },
    "ovh": {
        "report_url": "https://www.ovhcloud.com/en/abuse",
        "note": "OVH/OVHcloud — có form riêng.",
    },
    "azure": {
        "report_url": "https://msrc.microsoft.com/report/abuse",
        "note": "Microsoft Azure — report qua MSRC.",
    },
    "amazonaws": {
        "report_url": "https://support.aws.amazon.com/#/contacts/report-abuse",
        "note": "AWS — EC2/S3/CloudFront/ELB. Form Trust & Safety.",
    },
    "gcp": {
        "report_url": "https://support.google.com/code/contact/cloud_platform_report",
        "note": "Google Cloud Platform (GCP) — report qua form.",
    },
    "firebase": {
        "report_url": "https://support.google.com/legal/troubleshooter/1114905",
        "note": "Firebase Hosting (Google) — dùng form Legal Troubleshooter.",
    },
    "google-cloud": {
        "report_url": "https://support.google.com/code/contact/cloud_platform_report",
        "note": "Google Cloud Platform (App Engine, GCS...) — report qua form.",
    },
    "cloudflare-pages": {
        "report_url": "https://abuse.cloudflare.com",
        "note": "Cloudflare Pages (pages.dev) — dùng form abuse Cloudflare.",
    },
    "aws": {
        "report_url": "https://support.aws.amazon.com/#/contacts/report-abuse",
        "note": "AWS (S3, EC2, Elastic Beanstalk...) — form Trust & Safety.",
    },
    "render": {
        "report_url": "https://render.com/abuse",
        "note": "Render.com hosting — report qua form.",
    },
    "heroku": {
        "report_url": "https://www.heroku.com/policy/heroku-trust-and-safety-program",
        "note": "Heroku (Salesforce) — report qua trang Trust & Safety.",
    },
}

# Mail providers phổ biến hay bị phishing lợi dụng để gửi email lừa đảo.
# Khi domain phishing có MX record trỏ về 1 provider trong bảng này, cần report thêm
# tới abuse team của provider đó (không chỉ registrar/CA/CDN).
MAIL_PROVIDER_ABUSE = {
    "google":     {"name": "Google Workspace",  "abuse_url": "https://support.google.com/mail/contact/abuse",         "email": None},
    "googlemail": {"name": "Google Workspace",  "abuse_url": "https://support.google.com/mail/contact/abuse",         "email": None},
    "outlook":    {"name": "Microsoft 365",     "abuse_url": "https://msrc.microsoft.com/report/abuse",               "email": None},
    "office365":  {"name": "Microsoft 365",     "abuse_url": "https://msrc.microsoft.com/report/abuse",               "email": None},
    "microsoft":  {"name": "Microsoft 365",     "abuse_url": "https://msrc.microsoft.com/report/abuse",               "email": None},
    "zoho":       {"name": "Zoho Mail",         "abuse_url": "https://www.zoho.com/mail/spamreport.html",             "email": "abuse@zoho.com"},
    "yahoodns":   {"name": "Yahoo Mail",        "abuse_url": None,                                                    "email": "abuse@yahoo.com"},
    "protonmail": {"name": "ProtonMail",        "abuse_url": None,                                                    "email": "abuse@proton.me"},
    "mailgun":    {"name": "Mailgun",           "abuse_url": None,                                                    "email": "abuse@mailgun.com"},
    "sendgrid":   {"name": "SendGrid (Twilio)", "abuse_url": "https://sendgrid.com/policies/tos/",                    "email": "abuse@sendgrid.com"},
    "amazonaws":  {"name": "Amazon SES",        "abuse_url": "https://support.aws.amazon.com/#/contacts/report-abuse","email": "abuse@amazonaws.com"},
}

# Registry quản lý ccTLD — bảng tĩnh tra TRƯỚC (không cần mạng), chỉ TLD 1 nhãn (03_Technical_Guide.md
# mục 5 bước 3). TLD không có trong bảng này fallback sang iana_referral() (cần mạng).
CCTLD_REGISTRY_CONTACTS = {
    # --- ccTLD thật ---
    "cn": {
        "registry": "CNNIC",
        "abuse_email": "supervision@cnnic.cn",
        "note": "Yêu cầu bản dịch nhãn hiệu tiếng Trung/Anh có công chứng.",
    },
    "in": {"registry": "NIXI", "abuse_email": "abuse@registry.in", "note": "Hay không hồi âm — nên đi đường nhà đăng ký."},
    "io": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": None},
    "jp": {
        "registry": "JPRS",
        "abuse_email": "abuse@jprs.jp",
        "note": "Đòi hỏi xác thực nhãn hiệu đăng ký tại Nhật Bản.",
    },
    "kr": {
        "registry": "KISA",
        "abuse_email": "abuse@kisa.or.kr",
        "note": "Phối hợp báo cáo qua KrCERT.",
    },
    "ru": {
        "registry": "Coordination Center for TLD RU",
        "abuse_email": "abuse@cctld.ru",
        "note": None,
    },
    "uk": {
        "registry": "Nominet",
        "abuse_email": None,
        "report_webform": "https://registrars.nominet.uk/abuse-complaints/",
        "note": "Áp dụng cho cả .co.uk và .org.uk.",
    },
    "eu": {"registry": "EURid", "abuse_email": "info@eurid.eu", "note": None},
    "tw": {
        "registry": "TWNIC",
        "abuse_email": None,
        "report_webform": "https://phishingcheck.tw",
        "note": "Áp dụng cho cả .com.tw.",
    },
    "hk": {"registry": "HKIRC", "abuse_email": "abuse@hkirc.hk", "note": None},
    "de": {
        "registry": "DENIC",
        "abuse_email": None,
        "report_webform": "https://www.denic.de/en/report-a-concern/",
        "note": None,
    },
    "fr": {"registry": "AFNIC", "abuse_email": None, "note": None},
    "nl": {"registry": "SIDN", "abuse_email": "support@sidn.nl", "note": None},
    "pl": {"registry": "NASK", "abuse_email": "info@dns.pl", "note": None},
    "it": {"registry": "IIT-CNR", "abuse_email": "info@nic.it", "note": None},
    "mx": {"registry": "NIC Mexico", "abuse_email": "legal@nic.mx", "note": "Áp dụng cho cả .com.mx"},
    "au": {"registry": "auDA", "abuse_email": "compliance@auda.org.au", "note": None},
    "ca": {"registry": "CIRA", "abuse_email": "cira@cira.ca", "note": None},
    "br": {"registry": "NIC.br", "abuse_email": "abuse@registro.br", "note": None},
    "vn": {
        "registry": "VNNIC",
        "abuse_email": "abuse@vnnic.vn",
        "note": "Đối với domain .vn phishing, report thêm tới VNCERT (cert@vncert.gov.vn).",
    },
    # --- gTLD / nTLD phổ biến trong phishing ---
    "com": {
        "registry": "Verisign",
        "abuse_email": None,
        "report_webform": "https://verisign.my.site.com/DNSAbuse",
        "note": "Áp dụng cho cả .net .name .cc",
    },
    "net": {
        "registry": "Verisign",
        "abuse_email": None,
        "report_webform": "https://verisign.my.site.com/DNSAbuse",
        "note": None,
    },
    "name": {
        "registry": "Verisign",
        "abuse_email": None,
        "report_webform": "https://verisign.my.site.com/DNSAbuse",
        "note": None,
    },
    "cc": {
        "registry": "Verisign",
        "abuse_email": None,
        "report_webform": "https://verisign.my.site.com/DNSAbuse",
        "note": None,
    },
    "us": {
        "registry": "GoDaddy Registry",
        "abuse_email": None,
        "report_webform": "https://domainabuse.registry.godaddy/s/",
        "note": "Áp dụng cho cả .co .biz .vip .club (GoDaddy Registry).",
    },
    "co": {
        "registry": "Registry Co",
        "abuse_email": "abuse@registry.co",
        "report_webform": "https://registry.co/#contact",
        "note": None,
    },
    "biz": {
        "registry": "GoDaddy Registry",
        "abuse_email": None,
        "report_webform": "https://domainabuse.registry.godaddy/s/",
        "note": None,
    },
    "vip": {
        "registry": "GoDaddy Registry",
        "abuse_email": None,
        "report_webform": "https://domainabuse.registry.godaddy/s/",
        "note": None,
    },
    "club": {
        "registry": "GoDaddy Registry",
        "abuse_email": None,
        "report_webform": "https://domainabuse.registry.godaddy/s/",
        "note": None,
    },
    "top": {
        "registry": "Jiangsu Bangning Science & Technology (nic.top)",
        "abuse_email": "abuse@nic.top",
        "report_webform": "https://www.nic.top/cn/Complaintsnew.asp",
        "note": "Có xử lý thật — nhận email + web form. Gửi kèm file Excel (tối đa 200 domain/lần) và ảnh chụp màn hình có thanh địa chỉ.",
    },
    "xyz": {
        "registry": "XYZ.COM LLC",
        "abuse_email": "xyz_abuse@gen.xyz",
        "report_webform": "https://gen.xyz/account/abuse.php",
        "note": "Áp dụng cho 35+ TLD của XYZ (monster, quest, baby, cars, beauty, hair, homes, game, lol, mom, pics, hosting, audio, diet, ceo...). Nhiều domain cùng lúc gửi email xyz_abuse@gen.xyz. Hiệu quả — thường hôm sau bị khoá.",
    },
    # XYZ TLD family — cùng registry/abuse contact với .xyz
    "monster":  {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "quest":    {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "baby":     {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "cars":     {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "beauty":   {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "hair":     {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "homes":    {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "game":     {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "lol":      {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "mom":      {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "pics":     {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "hosting":  {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "audio":    {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "diet":     {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    "ceo":      {"registry": "XYZ.COM LLC", "abuse_email": "xyz_abuse@gen.xyz", "report_webform": "https://gen.xyz/account/abuse.php", "note": None},
    # CentralNic 2-level domains — không phải ccTLD thật, dùng report_webform chung
    "uk.com":   {"registry": "CentralNic", "abuse_email": None, "report_webform": "https://www.centralnicregistry.com/contact", "note": "Chọn mục Report Abuse. Không phải ccTLD UK — đừng gửi Nominet."},
    "jpn.com":  {"registry": "CentralNic", "abuse_email": None, "report_webform": "https://www.centralnicregistry.com/contact", "note": "Không phải ccTLD JP — đừng gửi JPRS."},
    "jp.net":   {"registry": "CentralNic", "abuse_email": None, "report_webform": "https://www.centralnicregistry.com/contact", "note": None},
    "in.net":   {"registry": "CentralNic", "abuse_email": None, "report_webform": "https://www.centralnicregistry.com/contact", "note": None},
    "com.co":   {"registry": "CentralNic", "abuse_email": None, "report_webform": "https://www.centralnicregistry.com/contact", "note": None},
    # Alias 2 cấp — trỏ về registry của TLD cha
    "co.in":    {"registry": "NIXI", "abuse_email": "abuse@registry.in", "note": "Hay không hồi âm — nên đi đường nhà đăng ký."},
    "co.jp":    {"registry": "JPRS", "abuse_email": "abuse@jprs.jp", "note": "Đòi hỏi xác thực nhãn hiệu đăng ký tại Nhật Bản."},
    "com.mx":   {"registry": "NIC Mexico", "abuse_email": "legal@nic.mx", "note": None},
    "shop": {
        "registry": "GMO Registry / Shopify",
        "abuse_email": None,
        "report_webform": "https://get.shop/abuse",
        "note": None,
    },
    "buzz": {"registry": "DOTSTRATEGY CO.", "abuse_email": "buzzabuse@dotstrategy.biz", "note": None},
    "site": {"registry": "Radix Registry", "abuse_email": "abuse@radix.support", "note": None},
    "online": {"registry": "Radix Registry", "abuse_email": "abuse@radix.support", "note": None},
    "store": {"registry": "Radix Registry", "abuse_email": "abuse@radix.support", "note": None},
    "pw": {"registry": "Radix Registry", "abuse_email": "abuse@radix.support", "note": "Formerly Palau ccTLD, now managed by Radix."},
    "live": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": "Web form có trên site họ nhưng không hoạt động — chỉ dùng email."},
    "ltd": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": None},
    "ninja": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": None},
    "pro": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": None},
    "vc": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": None},
    "info": {"registry": "Identity Digital", "abuse_email": "abuse@identity.digital", "note": None},
    "me": {"registry": "doMEn", "abuse_email": None, "report_webform": None, "note": "Không có kênh registry riêng — báo nhà đăng ký."},
    "tv": {"registry": "Verisign", "abuse_email": None, "report_webform": None, "note": "Không có kênh registry riêng — báo nhà đăng ký."},
}

# Bảng tĩnh abuse email của các registrar phổ biến — dùng làm fallback khi WHOIS không trả về email.
# Match theo "contains" case-insensitive với tên registrar từ WHOIS (xem lookup_registrar_abuse_email).
REGISTRAR_ABUSE_EMAILS = {
    "namecheap":           "abuse@namecheap.com",
    "godaddy":             "abuse@godaddy.com",
    "tucows":              "domainabuse@tucows.com",
    "opensrs":             "domainabuse@tucows.com",
    "enom":                "abuse@enom.com",
    "porkbun":             "abuse@porkbun.com",
    "gandi":               "abuse@support.gandi.net",
    "ovh":                 "abuse@ovh.net",
    "cloudflare":          "registrar-abuse@cloudflare.com",
    "google domains":      "registrar-abuse@google.com",
    "squarespace":         "registrar-abuse@google.com",
    "markmonitor":         "abusecomplaints@markmonitor.com",
    "alibaba":             "DomainAbuse@service.aliyun.com",
    "aliyun":              "DomainAbuse@service.aliyun.com",
    "huaweicloud":         "abuse@huaweicloud.com",
    "reg.ru":              "abuse@reg.ru",
    "namesilo":            "abuse@namesilo.com",
    "hostinger":           "abuse@hostinger.com",
    "hostgator":           "abuse@hostgator.com",
    "bluehost":            "abuse@bluehost.com",
    "dynadot":             "abuse@dynadot.com",
    "1api":                "abuse@1api.net",
    "internetbs":          "abuse@internet.bs",
    "cheapnames":          "abuse@cheapnames.com",
    "bigrock":             "abuse@bigrock.com",
    "crazydomains":        "abuse@crazydomains.com",
    "name.com":            "abuse@name.com",
    "rebel.com":           "abuse@rebel.com",
    "ionos":               "abuse@ionos.com",
    "1&1":                 "abuse@ionos.com",
    "netim":               "abuse@netim.com",
    "spaceship":           "abuse@spaceship.com",
    "epik":                "abuse@epik.com",
    "key-systems":         "abuse@key-systems.net",
    "inwx":                "abuse@inwx.com",
    "regtons":             "abuse@regtons.com",
    "publicdomainregistry": "abuse@publicdomainregistry.com",
    # Các entry dưới đây là WHOIS privacy services, KHÔNG phải abuse handlers thật.
    # Chỉ giữ để nhận diện, filter sẽ loại bỏ chúng trong generate_email_drafts().
    # "withheldforprivacy":  "abuse@withheldforprivacy.com",  # WHOIS privacy của Namecheap, không xử lý report
    # "domainsbyproxy":      "abuse@domainsbyproxy.com",      # WHOIS privacy của GoDaddy, không xử lý report
    "whoisguard":          "compliance@namecheap.com",
    "resellerclub":        "abuse-staff@resellerclub.com",
    "netearth":            "abuse@netearth.net",
    "nicenic":             "abuse@nicenic.net",
    "west263":             "abuse@hkdns.hk",
    "hkdns":               "abuse@hkdns.hk",
    "net-chinese":         "abuse@net-chinese.com.tw",
    "hichina":             "DomainAbuse@service.aliyun.com",
    "wix":                 "abuse@wix.com",
    "wordpress":           "abuse@automattic.com",
    "shopify":             "trust@shopify.com",
}

# Domain của các WHOIS privacy service — email từ các domain này KHÔNG phải abuse contact thật.
# Khi WHOIS trả về email thuộc 1 domain trong set này, bỏ qua và dùng fallback static table.
WHOIS_PRIVACY_DOMAINS = {
    "withheldforprivacy.com",   # Namecheap WHOIS privacy
    "domainsbyproxy.com",       # GoDaddy WHOIS privacy
    "domain-contact.org",       # WHOIS relay không xử lý report
    "privacyprotect.org",       # WHOIS privacy service
    "whoisguard.com",           # Namecheap WhoisGuard
    "contactprivacy.com",       # Contact Privacy Inc (Tucows)
    "anonymize.com",            # anonymize WHOIS
    "privacydaemon.com",        # Privacy Daemon
}

# Registrars CHỈ nhận report qua web form, KHÔNG nhận email.
# Khi registrar khớp 1 key trong dict này, sinh file hướng dẫn web form thay vì email.
WEB_FORM_REGISTRARS = {
    "godaddy":              "https://supportcenter.godaddy.com/abusereport/phishing",
    "dynadot":              "https://www.dynadot.com/report-abuse",
    "tucows":               "https://www.tucowsdomains.com/report-abuse",
    "opensrs":              "https://www.tucowsdomains.com/report-abuse",
    "namesilo":             "https://www.namesilo.com/report_abuse.php",
    "porkbun":              "https://www.porkbun.com/abuse",
    "sav.com":              "https://abuse.sav.com",
    "key-systems":          "https://abuse.cleandns.space",
    "instra":               "https://abuse.cleandns.space",
    "cleandns":             "https://abuse.cleandns.space",
    "gname":                "https://www.gname.com/abuse",
    "cosmotown":            "https://portal.cosmotown.com/report",
    "realtime register":    "https://www.mydomainprovider.com/contact_abuse",
    "publicdomainregistry": "https://www.publicdomainregistry.com/process-for-handling-abuse",
    "alibaba":              "https://report.alibabacloud.com",
    "aliyun":               "https://report.alibabacloud.com",
    "hichina":              "https://report.alibabacloud.com",
    "dnspod":               "https://intl.cloud.tencent.com/report-platform/dnsabuse",
    "tencent":              "https://intl.cloud.tencent.com/report-platform/dnsabuse",
    "west263":              "https://abuse.west.cn/",
    "verisign":             "https://verisign.my.site.com/DNSAbuse",
    "network solutions":    "https://www.networksolutions.com/support/aup/",
    "web.com":              "https://www.web.com/legal/abuse",
}


def get_webform_draft_text(
    domain: str,
    registrar: str,
    webform_url: str,
    cfg: dict,
    target_url: str = "",
    vt_link: str = "",
    urlscan: dict | None = None,
) -> str:
    """Sinh nội dung tố cáo để paste thẳng vào ô Description/Comments của web form.

    Không có field labels hay hướng dẫn — chỉ là đoạn text evidence sạch.
    urlscan: dict trả về từ urlscan_submit_and_wait() — nếu có thì thêm screenshot/result URL.
    """
    from datetime import datetime, timezone as _tz
    brand = cfg.get("brand_name") or "[BRAND]"
    contact_name = cfg.get("contact_name") or "[TÊN BẠN]"
    contact_email = cfg.get("contact_email") or "[EMAIL BẠN]"
    detected_date = datetime.now(_tz.utc).strftime("%Y-%m-%d")
    t_url = target_url or f"https://{domain}"
    vt = vt_link or f"https://www.virustotal.com/gui/domain/{domain}"

    evidence_lines = (
        f"The domain {domain} is being actively used for phishing — impersonating {brand} "
        f"to deceive users into submitting credentials and personal data.\n\n"
        f"Phishing URL: {t_url}\n"
        f"VirusTotal:   {vt}\n"
    )
    if urlscan and not urlscan.get("error"):
        if urlscan.get("result_url"):
            evidence_lines += f"URLScan:      {urlscan['result_url']}\n"
        if urlscan.get("screenshot_url"):
            evidence_lines += f"Screenshot:   {urlscan['screenshot_url']}\n"
    evidence_lines += (
        f"Detected:     {detected_date}\n\n"
        f"We request immediate suspension (serverHold / clientHold) of this domain.\n\n"
        f"Reported by: {contact_name} <{contact_email}>"
    )
    return evidence_lines


def lookup_registrar_abuse_email(registrar: str) -> str | None:
    """Tra bảng tĩnh REGISTRAR_ABUSE_EMAILS theo tên registrar (contains match, case-insensitive).
    Trả về abuse email nếu tìm được, None nếu không."""
    if not registrar:
        return None
    r_lower = registrar.lower()
    for key, email in REGISTRAR_ABUSE_EMAILS.items():
        if key in r_lower:
            return email
    return None


# RDAP bootstrap cache (TLD → RDAP server URL). Load 1 lần per process.
# RDAP là chuẩn mới của ICANN thay thế WHOIS text — trả về JSON có cấu trúc,
# có entity role "abuse" rõ ràng, đáng tin cậy hơn python-whois parse text nhiều.
_rdap_bootstrap_cache: dict = {}
_rdap_bootstrap_loaded: bool = False


def _load_rdap_bootstrap() -> dict:
    """Load IANA RDAP bootstrap cho DNS TLD (https://data.iana.org/rdap/dns.json).
    Cache sau lần load đầu — không fetch lại trong cùng 1 process."""
    global _rdap_bootstrap_cache, _rdap_bootstrap_loaded
    if _rdap_bootstrap_loaded:
        return _rdap_bootstrap_cache
    try:
        r = requests.get("https://data.iana.org/rdap/dns.json", timeout=8)
        if r.status_code == 200:
            data = r.json()
            mapping = {}
            for service in data.get("services", []):
                tlds, servers = service[0], service[1]
                if servers:
                    for tld in tlds:
                        mapping[tld.lower()] = servers[0].rstrip("/")
            _rdap_bootstrap_cache = mapping
    except Exception:
        pass
    _rdap_bootstrap_loaded = True
    return _rdap_bootstrap_cache


def get_rdap_abuse_email(domain: str) -> dict:
    """Tra abuse email của registrar qua RDAP (ICANN standard JSON, đáng tin hơn WHOIS text).

    Dùng IANA bootstrap để tìm RDAP server đúng cho TLD của domain — không hardcode,
    hoạt động cho mọi gTLD/ccTLD có hỗ trợ RDAP. Đây chính xác là cách ICANN Lookup
    (lookup.icann.org) tra thông tin abuse contact.

    Trả về dict với các key:
      "registrar"   : tên registrar (str) — nếu parse được
      "abuse_email" : email abuse (str) — nếu có
      "error"       : lý do thất bại (str) — nếu không lấy được

    KHÔNG raise exception — lỗi trả về trong key "error".
    """
    if requests is None:
        return {"error": "requests library not available"}

    tld = domain.lower().rsplit(".", 1)[-1]
    bootstrap = _load_rdap_bootstrap()
    rdap_base = bootstrap.get(tld)
    if not rdap_base:
        return {"error": f"No RDAP server for .{tld} in IANA bootstrap"}

    try:
        resp = requests.get(
            f"{rdap_base}/domain/{domain}",
            timeout=10,
            headers={"Accept": "application/rdap+json"},
        )
        if resp.status_code == 404:
            return {"error": "Domain not found in RDAP"}
        if resp.status_code != 200:
            return {"error": f"RDAP HTTP {resp.status_code}"}
        data = resp.json()
    except Exception as e:
        return {"error": str(e)}

    registrar_name = None
    abuse_email = None

    def _vcard_field(vcard_array, field_name):
        """Lấy giá trị 1 field từ vcardArray dạng [['fn',{},'text','Value'],...]."""
        if len(vcard_array) < 2:
            return None
        for prop in vcard_array[1]:
            try:
                if prop[0] == field_name:
                    return prop[3]
            except (IndexError, TypeError):
                continue
        return None

    for entity in data.get("entities", []):
        if "registrar" not in entity.get("roles", []):
            continue
        registrar_name = _vcard_field(entity.get("vcardArray", []), "fn") or registrar_name
        # Tìm entity con có role "abuse"
        for sub in entity.get("entities", []):
            if "abuse" in sub.get("roles", []):
                email = _vcard_field(sub.get("vcardArray", []), "email")
                if email:
                    abuse_email = email
                    break
        if registrar_name or abuse_email:
            break

    result = {}
    if registrar_name:
        result["registrar"] = registrar_name
    if abuse_email:
        result["abuse_email"] = abuse_email
    if not result:
        result["error"] = "No registrar/abuse entity found in RDAP response"
    return result




def load_config():
    cfg = configparser.ConfigParser()
    if os.path.exists(CONFIG_PATH):
        cfg.read(CONFIG_PATH, encoding="utf-8")

    # --- Multi-account SMTP (mới) + backward compat với cấu hình đơn cũ ---
    # Ưu tiên: nếu có key "accounts" (JSON array) → dùng đó.
    # Fallback: nếu chỉ có host/username/password cũ → wrap thành list 1 phần tử.
    accounts_raw = cfg.get("smtp", "accounts", fallback="").strip()
    if accounts_raw:
        try:
            smtp_accounts = json.loads(accounts_raw)
        except json.JSONDecodeError:
            smtp_accounts = []
    else:
        old_host = cfg.get("smtp", "host", fallback="")
        old_user = cfg.get("smtp", "username", fallback="")
        old_pass = cfg.get("smtp", "password", fallback="")
        if old_host and old_user and old_pass:
            smtp_accounts = [{
                "host": old_host,
                "port": cfg.getint("smtp", "port", fallback=587),
                "username": old_user,
                "password": old_pass,
            }]
        else:
            smtp_accounts = []

    proxies_raw = cfg.get("smtp", "proxies", fallback="").strip()
    if proxies_raw:
        try:
            smtp_proxies = json.loads(proxies_raw)
        except json.JSONDecodeError:
            smtp_proxies = []
    else:
        smtp_proxies = []

    return {
        "vt_api_key": cfg.get("api", "vt_api_key", fallback="") or os.environ.get("VT_API_KEY", ""),
        "gsb_api_key": cfg.get("api", "gsb_api_key", fallback="") or os.environ.get("GSB_API_KEY", ""),
        "urlscan_api_key": cfg.get("api", "urlscan_api_key", fallback="") or os.environ.get("URLSCAN_API_KEY", ""),
        "brand_name": cfg.get("company", "brand_name", fallback="[TÊN THƯƠNG HIỆU]"),
        "contact_name": cfg.get("company", "contact_name", fallback="[TÊN NGƯỜI BÁO CÁO]"),
        "contact_email": cfg.get("company", "contact_email", fallback="[EMAIL LIÊN HỆ]"),
        # Multi-account + proxy (mới)
        "smtp_accounts": smtp_accounts,
        "smtp_proxies": smtp_proxies,
        # Keys cũ giữ lại để backward compat với bất kỳ nơi nào còn dùng trực tiếp
        "smtp_host": cfg.get("smtp", "host", fallback=""),
        "smtp_port": cfg.getint("smtp", "port", fallback=587),
        "smtp_username": cfg.get("smtp", "username", fallback=""),
        "smtp_password": cfg.get("smtp", "password", fallback=""),
    }


# --------------------------------------------------------------------------
# Helpers dùng chung
# --------------------------------------------------------------------------

def normalize_domain(target: str) -> str:
    if "://" in target:
        target = urlparse(target).netloc
    return target.split(":")[0].split("/")[0].strip()


def get_cert_info(domain: str, port: int = 443, timeout: float = 8.0):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    with socket.create_connection((domain, port), timeout=timeout) as sock:
        ip = sock.getpeername()[0]
        with ctx.wrap_socket(sock, server_hostname=domain) as ssock:
            der_cert = ssock.getpeercert(binary_form=True)

    try:
        from cryptography import x509
        from cryptography.hazmat.backends import default_backend

        cert = x509.load_der_x509_certificate(der_cert, default_backend())
        return {
            "ip": ip,
            "issuer": cert.issuer.rfc4514_string(),
            "serial": format(cert.serial_number, "X"),
            "not_before": str(cert.not_valid_before_utc),
            "not_after": str(cert.not_valid_after_utc),
        }
    except ImportError:
        return {"ip": ip, "issuer": None, "serial": None}


# "TLD giả": các second-level domain được CentralNic/đối tác bán lại như thể TLD riêng
# (vd. đăng ký "chass.ru.com" thực chất là subdomain của ru.com, WHOIS thật nằm ở ru.com,
# không phải ở registry .com). Hay bị lợi dụng cho phishing vì rẻ, WHOIS lỏng lẻo hơn ccTLD thật.
FAKE_TLD_SUFFIXES = {
    "ru.com", "uk.com", "us.com", "eu.com", "gb.com", "gb.net", "de.com",
    "jpn.com", "sa.com", "se.com", "kr.com", "no.com", "qc.com", "br.com",
    "cn.com", "hu.com", "za.com", "uy.com",
}


def whois_query_domain(domain: str) -> str:
    parts = domain.lower().split(".")
    if len(parts) > 2 and ".".join(parts[-2:]) in FAKE_TLD_SUFFIXES:
        return ".".join(parts[-2:])
    return domain


def get_whois_info(domain: str):
    if whois is None:
        return {"error": "python-whois chưa được cài (pip install python-whois)"}
    try:
        w = whois.whois(whois_query_domain(domain))
        return {
            "registrar": w.get("registrar"),
            "creation_date": str(w.get("creation_date")),
            "expiration_date": str(w.get("expiration_date")),
            "name_servers": w.get("name_servers"),
            "emails": w.get("emails"),
            "status": w.get("status"),
        }
    except Exception as e:
        return {"error": str(e)}


def get_ip_whois(ip: str) -> dict:
    """Tra WHOIS của 1 địa chỉ IP (registry ARIN/RIPE/APNIC... qua RDAP) — dùng cho IP gốc tìm
    được qua scan_common_subdomains(), KHÁC hẳn get_whois_info() ở trên (domain WHOIS qua
    python-whois). 2 loại dữ liệu khác nhau, không thay thế nhau.

    depth=2 bắt buộc: mặc định (depth=1) RDAP chỉ trả về entity "registrant", không có entity
    role="abuse" — abuse contact thật của ISP/hosting chỉ xuất hiện khi tăng depth. contact.email
    trong RDAP là 1 list dict [{"value": "..."}], không phải string thẳng.
    """
    if IPWhois is None:
        return {"error": "ipwhois chưa được cài (pip install ipwhois)"}
    try:
        result = IPWhois(ip).lookup_rdap(depth=2)
        objects = result.get("objects") or {}

        abuse_email = None
        org = None
        for obj in objects.values():
            if "abuse" in (obj.get("roles") or []):
                emails = (obj.get("contact") or {}).get("email") or []
                if emails:
                    abuse_email = emails[0].get("value")
                break
        for obj in objects.values():
            # Ưu tiên tên tổ chức từ contact registrant (dễ đọc, vd "Google LLC") hơn
            # network.name (thường chỉ là handle ngắn, vd "GOGL") — dùng cho dòng "Dear {org}
            # Abuse Team" trong draft, cùng phong cách với generate_email_drafts().
            name = (obj.get("contact") or {}).get("name")
            if name:
                org = name
                break
        if not org:
            org = (result.get("network") or {}).get("name")

        return {
            "org": org,
            "abuse_email": abuse_email,
            "asn": result.get("asn"),
            "asn_description": result.get("asn_description"),
        }
    except Exception as e:
        return {"error": str(e)}


# --------------------------------------------------------------------------
# Registry WHOIS thô (raw socket) — leo thang ccTLD lạ (03_Technical_Guide.md mục 5)
# --------------------------------------------------------------------------
#
# Đây là loại WHOIS THỨ BA trong file này, khác cả 2 loại trên:
#   - get_whois_info()  → domain WHOIS qua thư viện python-whois (registrar/nameserver...)
#   - get_ip_whois()    → IP WHOIS qua RDAP/ipwhois (registry ARIN/RIPE/APNIC sở hữu dải IP)
#   - (dưới đây)        → nói chuyện TRỰC TIẾP với 1 WHOIS server bất kỳ qua giao thức WHOIS
#                         thô (raw socket port 43), vì python-whois tự chọn server theo TLD đã
#                         biết sẵn trong thư viện — không truy vấn được tới 1 WHOIS server tùy ý
#                         (cần cho ccTLD lạ/mới mà thư viện chưa biết, đi qua referral của IANA).

def query_whois_server(server: str, query: str, timeout: float = 8.0) -> str:
    """Gửi 1 truy vấn WHOIS thô tới `server` qua TCP port 43, trả về response dạng text.

    Không raise — lỗi kết nối/timeout trả về chuỗi rỗng, để lookup_registry_contact() coi như
    "không tra được" thay vì làm hỏng cả run_check(). Giống get_cert_info() ở trên: dùng
    socket.create_connection() thô, không qua thư viện WHOIS cấp cao (vì cần gửi query tới đúng
    server chỉ định, python-whois không hỗ trợ việc này). decode utf-8 với errors="ignore" vì
    WHOIS server của registry nước ngoài có thể trả về ký tự khó decode.
    """
    try:
        with socket.create_connection((server, 43), timeout=timeout) as sock:
            sock.sendall(f"{query}\r\n".encode())
            chunks = []
            while True:
                data = sock.recv(4096)
                if not data:
                    break
                chunks.append(data)
        return b"".join(chunks).decode("utf-8", errors="ignore")
    except Exception:
        return ""


_IANA_REFERRAL_RE = re.compile(r"^(refer|whois):\s*(\S+)", re.IGNORECASE | re.MULTILINE)


def iana_referral(tld: str):
    """Hỏi whois.iana.org xem TLD này do WHOIS server nào của registry phụ trách, trả về
    hostname WHOIS server đó (vd "whois.jprs.jp") hoặc None nếu không parse được.

    Ưu tiên dòng "whois:" hơn "refer:" nếu cả 2 cùng xuất hiện — "whois:" trỏ thẳng tới server
    của registry, còn "refer:" đôi khi chỉ trỏ tới 1 tầng trung gian khác.
    """
    raw = query_whois_server("whois.iana.org", tld)
    if not raw:
        return None
    fields = {key.lower(): value for key, value in _IANA_REFERRAL_RE.findall(raw)}
    return fields.get("whois") or fields.get("refer")



# whois.iana.org trả về referral cho MỌI TLD nó biết, kể cả gTLD — nếu không loại trừ,
# lookup_registry_contact() sẽ "leo thang Registry" cho cả domain .org/.mobi bình thường.
# .com/.net/.name/.cc đã được thêm vào CCTLD_REGISTRY_CONTACTS với Verisign web form (xem trên)
# nên không cần trong skip list nữa.
_SKIP_REGISTRY_ESCALATION_TLDS = {"org", "info", "biz", "name", "pro", "mobi"}


def lookup_registry_contact(domain: str) -> dict:
    """Tra abuse contact của Registry quản lý TLD của domain.

    Tra CCTLD_REGISTRY_CONTACTS (bảng tĩnh, không cần mạng) TRƯỚC — thử 2-level suffix
    (.co.jp, .uk.com, .com.mx...) trước, rồi mới thử 1-level (.jp, .com...).
    Fallback sang iana_referral() (cần mạng) khi không tìm thấy trong bảng.
    """
    d = domain.lower()
    # Thử 2-level suffix trước (vd "co.jp", "uk.com")
    parts = d.rsplit(".", 2)
    if len(parts) >= 3:
        two_level = f"{parts[-2]}.{parts[-1]}"
        if two_level in CCTLD_REGISTRY_CONTACTS:
            return {"source": "static_table", **CCTLD_REGISTRY_CONTACTS[two_level]}
    # Thử 1-level TLD
    tld = d.rsplit(".", 1)[-1]
    if tld in CCTLD_REGISTRY_CONTACTS:
        return {"source": "static_table", **CCTLD_REGISTRY_CONTACTS[tld]}
    if tld in _SKIP_REGISTRY_ESCALATION_TLDS:
        return {"source": "not_found"}
    referral_server = iana_referral(tld)
    if referral_server:
        raw = query_whois_server(referral_server, domain)
        return {"source": "iana_referral", "whois_server": referral_server, "raw": raw}
    return {"source": "not_found"}


def is_cloudflare(ns_list):
    if not ns_list:
        return False
    ns_list = ns_list if isinstance(ns_list, list) else [ns_list]
    return any("cloudflare.com" in str(ns).lower() for ns in ns_list)


# Hậu tố CNAME quen thuộc của từng CDN — chỉ nameserver là không đủ vì (khác Cloudflare) khách hàng
# của Fastly/Akamai/CloudFront thường vẫn giữ nguyên nameserver của registrar, chỉ CNAME bản ghi A.
_CDN_CNAME_SUFFIXES = {
    "fastly.net": "fastly",
    "akamaiedge.net": "akamai",
    "akamai.net": "akamai",
    "cloudfront.net": "cloudfront",
    # Vercel CNAME
    "vercel.app": "vercel",
    "vercel-dns.com": "vercel",
    # Netlify CNAME
    "netlify.app": "netlify",
    "netlify.com": "netlify",
    # GitHub Pages
    "github.io": "github-pages",
    "github.com": "github-pages",
    # Firebase / GCP
    "web.app": "firebase",
    "firebaseapp.com": "firebase",
    "appspot.com": "google-cloud",
    # DigitalOcean
    "ondigitalocean.app": "digitalocean",
    "digitaloceanspaces.com": "digitalocean",
    # Azure
    "azurewebsites.net": "azure",
    "azurestaticapps.net": "azure",
    "cloudapp.azure.com": "azure",
    # AWS (non-CloudFront)
    "amazonaws.com": "aws",
    "elasticbeanstalk.com": "aws",
    # Render
    "onrender.com": "render",
    # Heroku
    "herokuapp.com": "heroku",
}

# Hosting platform phishing thường host trực tiếp — detect qua domain suffix (đáng tin nhất)
_HOSTING_DOMAIN_SUFFIXES = {
    ".vercel.app": "vercel",
    ".netlify.app": "netlify",
    ".github.io": "github-pages",
    ".pages.dev": "cloudflare-pages",
    ".web.app": "firebase",
    ".firebaseapp.com": "firebase",
    ".onrender.com": "render",
    ".herokuapp.com": "heroku",
    ".azurewebsites.net": "azure",
    ".azurestaticapps.net": "azure",
    ".ondigitalocean.app": "digitalocean",
    ".appspot.com": "google-cloud",
    ".amazonaws.com": "aws",
    ".s3-website": "aws",
}


def detect_cdn(domain: str) -> list:
    """Nhận diện CDN/hosting platform đang che hoặc host domain.

    3 nguồn tín hiệu (theo thứ tự độ tin cậy giảm dần):
    1. Domain suffix — chắc chắn nhất (vd .vercel.app = luôn là Vercel)
    2. CNAME chain — qua dnspython
    3. HTTP response headers — Server/Via/X-Cache/X-Served-By/X-Amz-Cf-Id

    KHÔNG gồm Cloudflare — xem is_cloudflare() (dùng nameserver riêng).
    Không raise — lỗi ở từng nguồn chỉ khiến nguồn đó bị bỏ qua.
    """
    detected = set()
    domain_lower = domain.lower()

    # 1. Domain suffix check (ưu tiên cao nhất — không cần mạng)
    for suffix, name in _HOSTING_DOMAIN_SUFFIXES.items():
        if domain_lower.endswith(suffix):
            detected.add(name)

    # 2. CNAME chain
    if dns is not None:
        try:
            answer = dns.resolver.resolve(domain, "CNAME")
            for rdata in answer:
                target = str(rdata.target).rstrip(".").lower()
                for suffix, name in _CDN_CNAME_SUFFIXES.items():
                    if target.endswith(suffix):
                        detected.add(name)
        except Exception:
            pass

    # 3. HTTP headers
    if requests is not None:
        for scheme in ("http", "https"):
            try:
                r = requests.head(
                    f"{scheme}://{domain}/", timeout=5, allow_redirects=False, verify=False
                )
                blob = " ".join(
                    str(r.headers.get(h, ""))
                    for h in ("Server", "Via", "X-Cache", "X-Served-By", "X-Amz-Cf-Id")
                ).lower()
                powered_by = str(r.headers.get("X-Powered-By", "")).lower()
                if blob.strip() or powered_by:
                    if "cloudfront" in blob or r.headers.get("X-Amz-Cf-Id"):
                        detected.add("cloudfront")
                    if "fastly" in blob or "varnish" in blob:
                        detected.add("fastly")
                    if "akamai" in blob:
                        detected.add("akamai")
                    if "ddos-guard" in blob:
                        detected.add("ddos-guard")
                    if "stormwall" in blob:
                        detected.add("stormwall")
                    if "vercel" in blob or "vercel" in powered_by:
                        detected.add("vercel")
                    if "netlify" in blob or "netlify" in powered_by:
                        detected.add("netlify")
                    if "github" in blob:
                        detected.add("github-pages")
                    if "firebase" in blob or "firebasehosting" in blob:
                        detected.add("firebase")
                    if "heroku" in blob:
                        detected.add("heroku")
            except Exception:
                continue

    return sorted(detected)


# Wordlist subdomain thường gặp để dò IP gốc phía sau CDN/proxy (03_Technical_Guide.md mục 2).
COMMON_SUBDOMAINS = ["mail", "cpanel", "direct", "ftp", "dev", "staging", "webmail", "secure", "panel"]


def _resolve_subdomain(fqdn: str, timeout: float):
    try:
        if dns is not None:
            answer = dns.resolver.resolve(fqdn, "A", lifetime=timeout)
            return str(answer[0])
        old_timeout = socket.getdefaulttimeout()
        socket.setdefaulttimeout(timeout)
        try:
            return socket.gethostbyname(fqdn)
        finally:
            socket.setdefaulttimeout(old_timeout)
    except Exception:
        return None


def scan_common_subdomains(domain: str, timeout: float = 5.0) -> dict:
    """Resolve 1 wordlist subdomain cố định, trả về {subdomain: ip hoặc None}.

    1 subdomain lỗi/không resolve được không được làm hỏng cả vòng quét — mỗi lượt resolve
    (`_resolve_subdomain`) tự bắt exception riêng. Việc so sánh IP này với IP chính của domain
    (cert["ip"]) để tìm "candidate origin IP" được làm ở nơi hiển thị (cmd_check/Streamlit),
    không phải ở đây, vì hàm này không có quyền truy cập cert — giữ hàm thuần túy chỉ resolve DNS.

    Chạy SONG SONG (ThreadPoolExecutor) thay vì tuần tự — dùng dns.resolver (dnspython, có
    "lifetime" giới hạn timeout thật) khi có, thay vì socket.gethostbyname:
    socket.setdefaulttimeout() KHÔNG áp dụng được cho gethostbyname() (giới hạn đã biết của
    Python: timeout đó chỉ áp dụng cho socket connect/send/recv, không áp dụng cho resolve DNS
    qua thư viện hệ thống). Phát hiện lúc test: quét tuần tự 9 subdomain bằng gethostbyname mất
    77s dù "timeout" được set = 3.0; đổi sang dns.resolver(lifetime=...) sửa được việc timeout
    không có tác dụng, nhưng vẫn chậm (27-36s) vì DNS trong môi trường test có độ trễ thật
    ~4-5s/query — cộng dồn qua 9 subdomain tuần tự là không chấp nhận được cho 1 tool "check
    nhanh". Chạy song song thay vì tuần tự khiến tổng thời gian bị chặn bởi lookup CHẬM NHẤT
    (~timeout giây) thay vì TỔNG của cả 9 lookup — mới thực sự đạt được ý định "timeout ngắn"
    ban đầu của yêu cầu. Vẫn giữ socket.gethostbyname làm fallback nếu dnspython không có sẵn.
    """
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(COMMON_SUBDOMAINS)) as pool:
        future_to_sub = {
            pool.submit(_resolve_subdomain, f"{sub}.{domain}", timeout): sub
            for sub in COMMON_SUBDOMAINS
        }
        results = {sub: None for sub in COMMON_SUBDOMAINS}
        for future in concurrent.futures.as_completed(future_to_sub):
            results[future_to_sub[future]] = future.result()
    return results


def match_ca_notes(issuer: str):
    if not issuer:
        return None
    issuer_lower = issuer.lower()
    for key, info in CA_ABUSE_NOTES.items():
        if key in issuer_lower:
            return {"ca": key, **info}
    return None


# --------------------------------------------------------------------------
# VirusTotal
# --------------------------------------------------------------------------

def check_virustotal(domain: str, api_key: str):
    if not api_key:
        return {"skipped": "Chưa có VT API key trong config.ini"}
    if requests is None:
        return {"error": "Thư viện requests chưa được cài"}
    try:
        r = requests.get(
            f"https://www.virustotal.com/api/v3/domains/{domain}",
            headers={"x-apikey": api_key},
            timeout=15,
        )
        if r.status_code == 404:
            return {"note": "Domain chưa từng được VirusTotal quét trước đây"}
        r.raise_for_status()
        data = r.json()
        stats = data["data"]["attributes"]["last_analysis_stats"]
        return {
            "malicious": stats.get("malicious", 0),
            "suspicious": stats.get("suspicious", 0),
            "harmless": stats.get("harmless", 0),
            "total_engines": sum(stats.values()),
            "link": f"https://www.virustotal.com/gui/domain/{domain}",
        }
    except Exception as e:
        return {"error": str(e)}


def submit_virustotal(domain: str, api_key: str):
    """Gửi domain lên VirusTotal để quét nếu chưa có dữ liệu."""
    if not api_key or requests is None:
        return None
    try:
        r = requests.post(
            "https://www.virustotal.com/api/v3/urls",
            headers={"x-apikey": api_key},
            data={"url": f"http://{domain}"},
            timeout=15,
        )
        r.raise_for_status()
        return "Đã submit domain lên VirusTotal để quét, kiểm tra lại sau vài phút."
    except Exception as e:
        return f"Lỗi khi submit VirusTotal: {e}"


# --------------------------------------------------------------------------
# Google Safe Browsing (chỉ CHECK trạng thái — việc report vẫn phải làm
# thủ công qua form https://safebrowsing.google.com/safebrowsing/report_phish/
# vì Google không có API công khai để submit report)
# --------------------------------------------------------------------------

def check_safebrowsing(url: str, api_key: str):
    if not api_key:
        return {"skipped": "Chưa có GSB API key trong config.ini"}
    if requests is None:
        return {"error": "Thư viện requests chưa được cài"}
    try:
        body = {
            "client": {"clientId": "phishing-toolkit", "clientVersion": "1.0"},
            "threatInfo": {
                "threatTypes": ["MALWARE", "SOCIAL_ENGINEERING", "UNWANTED_SOFTWARE"],
                "platformTypes": ["ANY_PLATFORM"],
                "threatEntryTypes": ["URL"],
                "threatEntries": [{"url": url}],
            },
        }
        r = requests.post(
            f"https://safebrowsing.googleapis.com/v4/threatMatches:find?key={api_key}",
            json=body,
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("matches"):
            return {"flagged": True, "matches": data["matches"]}
        return {"flagged": False}
    except Exception as e:
        return {"error": str(e)}


def generate_safebrowsing_report_text(domain: str, cfg: dict) -> str:
    """Sinh sẵn 1 đoạn mô tả ngắn để copy-paste vào ô mô tả của form report thủ công tại
    https://safebrowsing.google.com/safebrowsing/report_phish/ (Google không có API submit report,
    chỉ có API tra cứu qua check_safebrowsing() ở trên).

    KHÔNG ghi ra file trong reports/ như các generate_*_draft() khác — đây là text dán vào 1 ô
    textbox của form web, không phải nội dung email gửi qua mail client hay SMTP, nên không cần
    (và không nên) có file .txt riêng hay xuất hiện trong danh sách gửi email của
    email_send_ui.py. Chỉ trả về string để cmd_check/pages/1_Check_Domain.py tự hiển thị bằng
    st.code()/print() cho người dùng copy.
    """
    rng = _draft_rng(domain)
    brand = cfg.get("brand_name") or "our brand"
    variants = [
        f"The domain {domain} is actively impersonating {brand} by cloning its official login page to harvest user credentials and OTP tokens. Please add this phishing URL to your browser security filters to protect users.",
        f"This site ({domain}) is a phishing page impersonating {brand}. It clones the official login interface to steal user credentials, OTP codes, and payment details. Please flag and block this URL.",
        f"{domain} is a phishing domain targeting {brand} users. The site replicates our official authentication portal to harvest login credentials and sensitive personal data.",
        f"We are reporting {domain} as an active phishing site impersonating {brand}. The page is designed to deceive users into submitting their login credentials, OTP tokens, and financial information.",
        f"This phishing URL ({domain}) fraudulently impersonates {brand}'s official website to steal user credentials and OTP tokens from victims. Immediate blocking is requested.",
    ]
    return _pick(rng, variants)


def generate_cloudflare_report_text(domain: str, cfg: dict) -> str:
    """Sinh sẵn nội dung mô tả để copy-paste vào form report Cloudflare tại
    https://abuse.cloudflare.com/ (mục Phishing & Malware) — tương tự
    generate_safebrowsing_report_text(), không ghi file, chỉ trả về string.
    """
    rng = _draft_rng(domain + "_cf")  # seed khác GSB để variant không trùng
    brand = cfg.get("brand_name") or "our brand"
    variants = [
        f"The domain {domain} is impersonating {brand} by cloning its official login page to harvest user credentials, OTP tokens, and payment details. This domain is using Cloudflare services to proxy and conceal the phishing infrastructure. Please suspend Cloudflare services for this domain and/or display a phishing interstitial warning.",
        f"{domain} is an active phishing site that impersonates {brand}'s official website to steal user credentials and financial data. The site is proxied through Cloudflare, which is being used to mask the true hosting origin. Please suspend Cloudflare access for this domain.",
        f"We are reporting {domain} as a phishing domain using Cloudflare to host and proxy a fraudulent site impersonating {brand}. The site harvests login credentials, OTP tokens, and payment information from users. Please terminate Cloudflare services for this account.",
        f"This domain ({domain}) clones {brand}'s login portal to steal credentials and sensitive data from victims. It leverages Cloudflare infrastructure to remain online and mask its origin. Immediate suspension of Cloudflare services for this domain is requested.",
    ]
    return _pick(rng, variants)



# --------------------------------------------------------------------------
# A1 — HTTP check: page còn sống không, title, redirect, login form
# --------------------------------------------------------------------------

def check_http(target: str) -> dict:
    """Fetch trang phishing để lấy HTTP status, page title, redirect chain và
    phát hiện login form (input type=password) mà không cần truy cập thủ công.

    Thử HTTPS trước, fallback HTTP nếu SSL error. verify=False vì domain nghi vấn
    thường có cert không hợp lệ. Chỉ đọc 100KB đầu để tránh download trang nặng.
    Kết quả không ghi vào case_log.csv (không thay đổi schema CSV đã có).
    """
    if requests is None:
        return {"error": "requests not installed"}

    ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    )
    if "://" in target:
        parsed = urlparse(target)
        urls = [target]
        if parsed.scheme.lower() == "https":
            urls.append(parsed._replace(scheme="http").geturl())
    else:
        urls = [f"https://{target}", f"http://{target}"]

    for url in urls:
        scheme = urlparse(url).scheme
        try:
            r = requests.get(
                url, timeout=10, allow_redirects=True, verify=False,
                headers={"User-Agent": ua}, stream=True,
            )
            # Giới hạn 100KB — tránh download trang nặng, đủ để parse title + form
            raw = b""
            for chunk in r.iter_content(8192):
                raw += chunk
                if len(raw) >= 100_000:
                    break
            text = raw.decode("utf-8", errors="replace")

            title_m = re.search(r"<title[^>]*>([^<]{1,200})", text, re.IGNORECASE)
            has_password = bool(re.search(r'type=["\']password["\']', text, re.IGNORECASE))
            form_count = len(re.findall(r"<form[\s>]", text, re.IGNORECASE))
            redirected = bool(r.history)
            final_url = r.url if r.url != url else None

            return {
                "status_code": r.status_code,
                "scheme": scheme,
                "page_title": title_m.group(1).strip() if title_m else None,
                "has_password_input": has_password,
                "form_count": form_count,
                "redirected": redirected,
                "final_url": final_url,
            }
        except requests.exceptions.SSLError:
            continue  # thử http
        except requests.exceptions.ConnectionError:
            return {"status_code": None, "error": "Không kết nối được — domain có thể đã down"}
        except Exception as e:
            if scheme == "http":
                return {"error": str(e)}
            continue

    return {"error": "Cả HTTPS và HTTP đều thất bại"}


# --------------------------------------------------------------------------
# A2 — Domain age: số ngày kể từ ngày đăng ký
# --------------------------------------------------------------------------

def compute_domain_age_days(who: dict):
    """Tính số ngày domain đã tồn tại từ creation_date của WHOIS.

    get_whois_info() trả về creation_date dạng str(w.get("creation_date")) — có thể là
    "2020-01-15 00:00:00", "[datetime.datetime(2020,1,15,0,0)]" (list repr), hoặc "None".
    Dùng regex tìm pattern YYYY-MM-DD đầu tiên trong chuỗi — đơn giản và đủ robust.
    Trả về int (số ngày) hoặc None nếu không parse được.
    """
    if not isinstance(who, dict):
        return None
    raw = str(who.get("creation_date") or "")
    if not raw or raw == "None":
        return None
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", raw)
    if not m:
        return None
    try:
        dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)), tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt).days
    except Exception:
        return None


# --------------------------------------------------------------------------
# B2 — MX record check: domain có gửi email không? Provider nào?
# --------------------------------------------------------------------------

def check_mx_records(domain: str) -> dict:
    """Resolve MX records của domain để xác định domain có khả năng gửi email không.

    Nếu có MX records → domain đang được cấu hình để gửi/nhận email → có thể đang
    dùng để gửi phishing email → cần report thêm tới abuse team của mail provider.
    Trả về {"records": [...], "providers": [...]} — providers là list dict từ
    MAIL_PROVIDER_ABUSE cho các provider được nhận diện qua hostname MX.
    Không raise — lỗi trả về {"records": [], "providers": [], "error": str(e)}.
    """
    if dns is None:
        # Fallback: không có dnspython, thử socket (không hỗ trợ timeout per-query)
        return {"records": [], "providers": [], "skipped": "dnspython không có sẵn — pip install dnspython"}

    records = []
    providers_found = {}  # key: provider_key, value: info dict (dedup)

    try:
        answers = dns.resolver.resolve(domain, "MX", lifetime=8.0)
        for rdata in sorted(answers, key=lambda r: r.preference):
            host = str(rdata.exchange).rstrip(".")
            records.append({"host": host, "priority": rdata.preference})
            host_lower = host.lower()
            for key, info in MAIL_PROVIDER_ABUSE.items():
                if key in host_lower and key not in providers_found:
                    providers_found[key] = info
    except dns.resolver.NoAnswer:
        return {"records": [], "providers": []}
    except dns.resolver.NXDOMAIN:
        return {"records": [], "providers": [], "note": "Domain không tồn tại (NXDOMAIN)"}
    except Exception as e:
        return {"records": [], "providers": [], "error": str(e)}

    return {
        "records": records,
        "providers": list(providers_found.values()),
    }


# --------------------------------------------------------------------------
# A3 — URLScan.io: screenshot + DOM analysis (free public API, no key needed)
# --------------------------------------------------------------------------

def urlscan_submit(domain: str, api_key: str = "", use_http: bool = False) -> dict:
    """Submit domain lên URLScan.io để lấy screenshot + DOM/DNS analysis.

    Cần API key miễn phí — đăng ký tại https://urlscan.io/ rồi điền urlscan_api_key
    vào config.ini. Rate limit free: 5000 scans/ngày.
    use_http=True để submit http:// thay vì https:// khi domain không có SSL hoặc bị
    URLScan từ chối do DNS/SSL error.
    """
    if requests is None:
        return {"error": "requests not installed"}
    if not api_key:
        return {"error": "Chưa có urlscan_api_key trong config.ini — đăng ký free tại https://urlscan.io/"}
    scheme = "http" if use_http else "https"
    try:
        r = requests.post(
            "https://urlscan.io/api/v1/scan/",
            json={"url": f"{scheme}://{domain}", "visibility": "public"},
            headers={"Content-Type": "application/json", "API-Key": api_key},
            timeout=15,
        )
        if r.status_code == 429:
            return {"error": "URLScan.io rate limit — thử lại sau vài phút"}
        if r.status_code in (400, 401, 403):
            try:
                msg = r.json().get("message", r.text)
            except Exception:
                msg = r.text
            # DNS Error = domain không resolve được từ phía URLScan — tức là domain đã down
            # hoặc chưa propagate DNS. Trả về warning thay vì error để UI hiện khác biệt.
            if "DNS" in msg or "resolve" in msg.lower() or "Could not" in msg:
                return {
                    "warning": f"URLScan không resolve được domain này ({msg}). "
                               "Có thể domain đã bị gỡ, DNS chưa propagate, hoặc chỉ hoạt động qua IPv6. "
                               "Thử dùng http:// thay vì https:// hoặc kiểm tra lại domain.",
                    "dns_error": True,
                }
            return {"error": f"URLScan từ chối ({r.status_code}): {msg}"}
        r.raise_for_status()
        data = r.json()
        uuid = data.get("uuid") or ""
        return {
            "scan_id": uuid,
            "result_url": f"https://urlscan.io/result/{uuid}/",
            "screenshot_url": f"https://urlscan.io/screenshots/{uuid}.png",
            "status": "submitted",
        }
    except Exception as e:
        return {"error": str(e)}


def urlscan_result(scan_uuid: str, api_key: str = "") -> dict:
    """Lấy kết quả scan đã submit lên URLScan.io. Cần truyền api_key giống lúc submit.

    Trả về {"status": "pending"} nếu scan chưa xong (HTTP 404).
    Trả về dict đầy đủ khi done: screenshot_url, malicious verdict, tags, page_title, IP, country.
    """
    if requests is None:
        return {"error": "requests not installed"}
    try:
        headers = {"API-Key": api_key} if api_key else {}
        r = requests.get(
            f"https://urlscan.io/api/v1/result/{scan_uuid}/",
            headers=headers,
            timeout=15,
        )
        if r.status_code == 404:
            return {"status": "pending"}
        r.raise_for_status()
        data = r.json()
        verdicts = ((data.get("verdicts") or {}).get("overall")) or {}
        page = data.get("page") or {}
        return {
            "status": "done",
            "screenshot_url": f"https://urlscan.io/screenshots/{scan_uuid}.png",
            "result_url": f"https://urlscan.io/result/{scan_uuid}/",
            "malicious": verdicts.get("malicious", False),
            "score": verdicts.get("score", 0),
            "tags": verdicts.get("tags") or [],
            "brands": verdicts.get("brands") or [],
            "page_title": page.get("title"),
            "page_ip": page.get("ip"),
            "page_country": page.get("country"),
        }
    except Exception as e:
        return {"error": str(e)}


def urlscan_submit_and_wait(domain: str, api_key: str, timeout: int = 65, poll_interval: int = 5) -> dict:
    """Submit URLScan và poll tự động đến khi có kết quả (hoặc hết timeout).

    Dùng để chạy song song với pipeline chính trong run_check() — submit ngay từ đầu,
    poll ở cuối khi các bước VT/GSB/WHOIS đã xong nên phần lớn thời gian 30s đã trôi qua.
    Trả về dict có status="done" nếu thành công, "pending"/error nếu timeout/lỗi.
    """
    import time as _time
    sub = urlscan_submit(domain, api_key)
    if "error" in sub or "warning" in sub:
        return sub
    scan_id = sub.get("scan_id")
    if not scan_id:
        return {"error": "URLScan: no scan_id returned"}
    deadline = _time.time() + timeout
    _time.sleep(poll_interval)  # đợi lần đầu trước khi bắt đầu poll
    while _time.time() < deadline:
        res = urlscan_result(scan_id, api_key)
        if res.get("status") == "done":
            return res
        if "error" in res:
            return res
        _time.sleep(poll_interval)
    return {"status": "pending", "scan_id": scan_id,
            "screenshot_url": f"https://urlscan.io/screenshots/{scan_id}.png",
            "result_url": f"https://urlscan.io/result/{scan_id}/"}


def append_urlscan_evidence_to_drafts(drafts: list, urlscan_res: dict) -> list:
    """Append URLScan.io evidence (screenshot link, result URL, verdict) vào cuối tất cả
    draft files đã sinh. Đồng thời thay thế dòng placeholder ảnh chụp màn hình trong
    draft bằng URL screenshot thật từ URLScan.

    Trả về list các draft path đã được cập nhật thành công.
    """
    if not drafts:
        return []
    # Cho phép chạy kể cả khi status="pending" (timeout) — ta vẫn có screenshot_url từ UUID
    screenshot_url = urlscan_res.get("screenshot_url", "")
    result_url = urlscan_res.get("result_url", "")

    if urlscan_res.get("status") == "done":
        verdict = "MALICIOUS" if urlscan_res.get("malicious") else "NOT flagged"
        score = urlscan_res.get("score", 0)
        tags = ", ".join(urlscan_res.get("tags") or []) or "—"
        brands = ", ".join(urlscan_res.get("brands") or []) or "—"
        page_title = urlscan_res.get("page_title") or "—"
        page_ip = urlscan_res.get("page_ip") or "—"
        country = urlscan_res.get("page_country") or "—"
        evidence_block = (
            f"\n\n--- Evidence: URLScan.io Analysis ---\n"
            f"Verdict   : {verdict} (score: {score})\n"
            f"Page title: {page_title}\n"
            f"Page IP   : {page_ip} ({country})\n"
            f"Tags      : {tags}\n"
            f"Brands    : {brands}\n"
            f"Screenshot: {screenshot_url}\n"
            f"Full report: {result_url}\n"
            f"--- End of URLScan evidence ---\n"
        )
    else:
        # Timeout hoặc pending — chỉ gắn URL, không có verdict
        evidence_block = (
            f"\n\n--- Evidence: URLScan.io (kết quả đang xử lý) ---\n"
            f"Screenshot: {screenshot_url}\n"
            f"Full report: {result_url}\n"
            f"--- End of URLScan evidence ---\n"
        ) if screenshot_url else ""

    updated = []
    for path in drafts:
        try:
            if not os.path.isfile(path):
                continue
            with open(path, encoding="utf-8") as f:
                content = f.read()
            # Không xử lý nếu đã có rồi (tránh duplicate)
            if "URLScan.io" in content and "urlscan.io/screenshots" in content:
                updated.append(path)
                continue
            # Thay thế placeholder ảnh chụp màn hình bằng link thật nếu có screenshot_url,
            # đồng thời xóa dòng instruction "(Ảnh chụp phải hiển thị rõ URL ...)" bên dưới
            if screenshot_url:
                import re as _re
                content = content.replace(
                    "[ĐÍNH KÈM ẢNH CHỤP MÀN HÌNH — bắt buộc có thanh địa chỉ trình duyệt]",
                    screenshot_url,
                )
                # Xóa dòng instruction động bên dưới (dòng này chứa URL trang phishing)
                content = _re.sub(
                    r'\(Ảnh chụp phải hiển thị rõ URL ".*?" trên thanh địa chỉ của trình duyệt\)\n?',
                    "",
                    content,
                )
            new_content = content + evidence_block if evidence_block else content
            with open(path, "w", encoding="utf-8") as f:
                f.write(new_content)
            updated.append(path)
        except Exception:
            pass  # file bị khóa hoặc lỗi khác — bỏ qua, không làm crash UI
    return updated


# --------------------------------------------------------------------------
# B3 — Wayback Machine: archive trang phishing làm bằng chứng tĩnh
# --------------------------------------------------------------------------

def wayback_archive(domain: str) -> dict:
    """Archive domain lên Wayback Machine làm bằng chứng tĩnh.

    Quy trình 2 bước:
    1. Kiểm tra snapshot cũ qua Availability API — nếu đã có snapshot trong 7 ngày gần nhất,
       trả về ngay mà không cần archive lại (tránh rate limit + nhanh hơn).
    2. Nếu chưa có, gửi Save Page Now (GET /save/{url}) và đọc Content-Location header.
       Nhiều error code từ Wayback (520, 523...) là lỗi upstream khi Wayback không thể
       fetch domain — thường gặp với domain mới/down/chặn bot. Xử lý gracefully, trả về
       link archive dự kiến (có thể verify thủ công) thay vì báo fail hoàn toàn.
    Không raise — lỗi trả về {"error": str(e), "manual_url": ...}.
    """
    if requests is None:
        return {"error": "requests not installed"}

    url = f"https://{domain}"
    manual_url = f"https://web.archive.org/web/*/{url}"  # link tra cứu thủ công luôn có sẵn

    # ── Bước 1: kiểm tra snapshot đã có qua Availability API (nhanh, không tốn quota) ──
    try:
        avail = requests.get(
            "https://archive.org/wayback/available",
            params={"url": url, "timestamp": datetime.now(timezone.utc).strftime("%Y%m%d")},
            timeout=10,
            headers={"User-Agent": "Mozilla/5.0 (compatible; PhishingToolkit/1.0)"},
        )
        if avail.status_code == 200:
            snap = avail.json().get("archived_snapshots", {}).get("closest", {})
            if snap.get("available") and snap.get("url"):
                snap_ts = snap.get("timestamp", "")
                # Chỉ dùng snapshot cũ nếu còn trong vòng 7 ngày
                try:
                    snap_dt = datetime.strptime(snap_ts[:8], "%Y%m%d").replace(tzinfo=timezone.utc)
                    age_days = (datetime.now(timezone.utc) - snap_dt).days
                    if age_days <= 7:
                        return {
                            "archive_url": snap["url"],
                            "status": "existing_snapshot",
                            "note": f"Snapshot có sẵn từ {snap_ts[:8]} (cách đây {age_days} ngày) — không cần archive lại.",
                        }
                except Exception:
                    pass
    except Exception:
        pass  # Availability API lỗi — tiếp tục thử save

    # ── Bước 2: gửi Save Page Now ────────────────────────────────────────────
    try:
        r = requests.get(
            f"https://web.archive.org/save/{url}",
            timeout=35,
            allow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (compatible; PhishingToolkit/1.0)"},
        )
        # Wayback trả về Content-Location header dạng /web/20240101120000/https://...
        loc = r.headers.get("Content-Location", "")
        if loc.startswith("/web/"):
            return {"archive_url": f"https://web.archive.org{loc}", "status": "archived"}
        if "web.archive.org/web/" in r.url:
            return {"archive_url": r.url, "status": "archived"}
        if r.status_code == 200:
            ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
            return {
                "archive_url": f"https://web.archive.org/web/{ts}/{url}",
                "status": "archived",
                "note": "Link ước tính — verify tại web.archive.org nếu cần.",
            }
        if r.status_code == 429:
            return {"error": "Wayback Machine rate limit — thử lại sau vài phút.", "manual_url": manual_url}
        # 520, 523, 5xx: Wayback không thể fetch domain (domain down, chặn bot, lỗi upstream)
        # Vẫn trả về link dự kiến để đính kèm vào report — người nhận có thể tự verify
        if r.status_code >= 500:
            ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
            return {
                "archive_url": f"https://web.archive.org/web/{ts}/{url}",
                "status": "save_failed",
                "note": (
                    f"Wayback trả về HTTP {r.status_code} khi cố archive — "
                    "domain có thể đang down, chặn bot, hoặc Wayback gặp lỗi upstream. "
                    "Link bên dưới là ước tính, cần verify thủ công."
                ),
                "manual_url": manual_url,
            }
        return {"error": f"Wayback trả về HTTP {r.status_code}", "manual_url": manual_url}
    except Exception as e:
        return {"error": str(e), "manual_url": manual_url}


def append_wayback_evidence_to_drafts(drafts: list, wayback_result: dict) -> list:
    """Append link Wayback Machine archive vào cuối tất cả draft files đã sinh.

    Idempotent: kiểm tra 'Wayback Machine Archive' đã có trong file chưa trước khi append.
    Cùng pattern với append_urlscan_evidence_to_drafts — không crash nếu file bị khóa.
    """
    archive_url = wayback_result.get("archive_url")
    if not drafts or not archive_url:
        return []

    note = wayback_result.get("note", "")
    evidence_block = (
        f"\n\n--- Evidence: Wayback Machine Archive ---\n"
        f"Archived URL : {archive_url}\n"
        + (f"Note         : {note}\n" if note else "")
        + f"--- End of Wayback evidence ---\n"
    )

    updated = []
    for path in drafts:
        try:
            if not os.path.isfile(path):
                continue
            with open(path, encoding="utf-8") as f:
                content = f.read()
            if "Wayback Machine Archive" in content:
                updated.append(path)
                continue
            with open(path, "a", encoding="utf-8") as f:
                f.write(evidence_block)
            updated.append(path)
        except Exception:
            pass
    return updated

def compute_reputation(vt: dict, gsb: dict) -> dict:
    vt_malicious = vt.get("malicious") or 0
    vt_suspicious = vt.get("suspicious") or 0
    gsb_flagged = gsb.get("flagged") is True

    reasons = []
    if vt_malicious:
        reasons.append(f"VirusTotal: {vt_malicious} engine gắn cờ malicious")
    if vt_suspicious:
        reasons.append(f"VirusTotal: {vt_suspicious} engine gắn cờ suspicious")
    if gsb_flagged:
        reasons.append("Google Safe Browsing: đã gắn cờ URL này")

    vt_unknown = "error" in vt or "skipped" in vt or "note" in vt
    gsb_unknown = "error" in gsb or "skipped" in gsb

    if vt_malicious or gsb_flagged:
        verdict = "flagged"
        label = "ĐÃ BỊ GẮN CỜ bởi bên thứ 3"
    elif vt_suspicious:
        verdict = "suspicious"
        label = "Có dấu hiệu nghi ngờ (suspicious)"
    elif vt_unknown and gsb_unknown:
        verdict = "unknown"
        label = "Chưa đủ dữ liệu để đánh giá (thiếu API key hoặc domain chưa từng bị quét)"
    else:
        verdict = "clean"
        label = "Chưa bên nào gắn cờ tại thời điểm kiểm tra"

    return {"verdict": verdict, "label": label, "reasons": reasons}


# --------------------------------------------------------------------------
# crt.sh — Certificate Transparency search (tìm domain "anh em")
# --------------------------------------------------------------------------

def crtsh_related(keyword: str):
    if requests is None:
        return {"error": "Thư viện requests chưa được cài"}
    try:
        r = requests.get(
            "https://crt.sh/", params={"q": f"%{keyword}%", "output": "json"}, timeout=25
        )
        r.raise_for_status()
        entries = r.json()
        names = set()
        for e in entries:
            for n in str(e.get("name_value", "")).split("\n"):
                n = n.strip().lower()
                if n and not n.startswith("*."):
                    names.add(n)
        return {"count": len(names), "domains": sorted(names)}
    except Exception as e:
        return {"error": str(e)}


# --------------------------------------------------------------------------
# dnstwist — brand monitoring / typosquat scan
# --------------------------------------------------------------------------

def brand_scan(domain: str, limit: int = 50):
    try:
        proc = subprocess.run(
            [sys.executable, "-m", "dnstwist", "-r", "-t", "60", "-f", "json", domain],
            capture_output=True,
            text=True,
            timeout=600,
        )
        if proc.returncode != 0 and not proc.stdout.strip():
            return {"error": proc.stderr.strip() or "dnstwist chạy lỗi (đã cài dnstwist chưa?)"}
        results = json.loads(proc.stdout)
        return {"count": len(results), "results": results[:limit]}
    except FileNotFoundError:
        return {"error": "Chưa cài dnstwist (pip install dnstwist)"}
    except subprocess.TimeoutExpired:
        return {"error": "dnstwist quét quá 10 phút chưa xong (domain/brand quá phổ biến, quá nhiều biến thể). Thử lại với --fuzzers để giới hạn bớt thuật toán sinh biến thể."}
    except Exception as e:
        return {"error": str(e)}


# --------------------------------------------------------------------------
# CSV log
# --------------------------------------------------------------------------

def log_case(row: dict):
    is_new = not os.path.exists(LOG_PATH)
    with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(row.keys()))
        if is_new:
            writer.writeheader()
        writer.writerow(row)


# --------------------------------------------------------------------------
# Email drafts
# --------------------------------------------------------------------------

# Seed random bằng domain + ngày để cùng 1 domain trong 1 ngày cho ra cùng variant
# nhưng ngày khác nhau thì khác → tránh gửi đúng 1 bản lặp lại nhiều lần.
def _draft_rng(domain: str) -> random.Random:
    seed = domain + datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return random.Random(seed)


def _pick(rng: random.Random, options: list) -> str:
    return rng.choice(options)


# Pool các variant cho từng phần của email — mỗi pool có ít nhất 4 variant.
_SUBJECT_REGISTRAR = [
    "Phishing Report - {domain} - Brand Impersonation",
    "Abuse Report: Active Phishing Site - {domain}",
    "Brand Impersonation Complaint - {domain}",
    "Phishing Domain Takedown Request - {domain}",
    "Formal Phishing Report - {domain}",
]

_OPENING_REGISTRAR = [
    "We are writing to report the domain {domain} for active phishing activity targeting users of our brand.",
    "We are reporting the domain {domain}, which is currently being used to conduct a phishing attack against our brand.",
    "This is a formal abuse report regarding {domain}, a domain actively engaged in phishing activity impersonating our brand.",
    "We have identified {domain} as a phishing domain actively targeting our users and brand identity.",
]

_DESCRIPTION_REGISTRAR = [
    "This domain is impersonating our brand's official website by cloning its login interface to harvest user credentials, OTP tokens, and personal data without authorization. Users who visit this site are deceived into submitting sensitive information (login credentials, passwords, payment details, etc.).",
    "The domain has been set up to mimic our official website in order to trick visitors into entering their credentials and personal data. It presents a convincing replica of our login page designed to steal user account information.",
    "This fraudulent site clones our brand's interface to deceive users into submitting sensitive data including login credentials, OTP tokens, and payment information. The phishing page is actively harvesting user data.",
    "The domain hosts a phishing page that impersonates our brand's login portal. Visitors are tricked into providing credentials, personal information, and financial data to the attacker.",
]

_REQUEST_REGISTRAR = [
    "We request the immediate suspension or transfer-lock of this domain in accordance with your Acceptable Use Policy and ICANN Registrar Accreditation Agreement (RAA) §3.18.",
    "We respectfully request that you suspend or place a transfer-lock on this domain immediately, pursuant to your Acceptable Use Policy and ICANN RAA §3.18.",
    "Please take immediate action to suspend this domain under your Acceptable Use Policy and ICANN Registrar Accreditation Agreement obligations (RAA §3.18).",
    "We urge you to suspend this domain without delay in compliance with your Acceptable Use Policy and ICANN RAA §3.18 requirements.",
]

_CLOSING_REGISTRAR = [
    "Please confirm receipt and any action taken at your earliest convenience.",
    "We appreciate your prompt attention to this matter and look forward to your confirmation.",
    "Your swift action on this matter is greatly appreciated. Please acknowledge receipt of this report.",
    "Thank you for your attention to this urgent matter. Please confirm once action has been taken.",
]

_SUBJECT_CA = [
    "SSL Certificate Revocation Request - {domain} (Phishing)",
    "Certificate Abuse Report: Phishing Domain - {domain}",
    "Fraudulent SSL Certificate in Use - {domain}",
    "SSL Certificate Revocation Request for Phishing Site - {domain}",
]

_OPENING_CA = [
    "We are requesting the immediate revocation of the SSL certificate issued for the domain {domain}, which is being used to impersonate our brand for phishing and credential harvesting.",
    "We are writing to report that the SSL certificate issued for {domain} is actively being used to facilitate a phishing attack against our brand.",
    "This is an urgent request for revocation of the SSL certificate for {domain}, a domain currently conducting phishing operations against our users.",
    "We have identified that the SSL certificate issued for {domain} is being misused to host a phishing site impersonating our brand.",
]

_DESCRIPTION_CA = [
    "The fraudulent site clones our brand's official login page to deceive users into submitting sensitive information (credentials, OTP tokens, payment details).",
    "This domain hosts a phishing page that replicates our brand's login interface to steal user credentials and personal information.",
    "The site mimics our official brand portal to harvest user login credentials, OTP tokens, and payment information from unsuspecting visitors.",
    "Attackers are using this certificate to add apparent legitimacy to a phishing page that clones our brand's authentication interface.",
]



def generate_email_drafts(domain, cert, who, vt, cfg, target_url=None):
    os.makedirs(REPORTS_DIR, exist_ok=True)
    written = []

    contact_name = cfg.get("contact_name") or ""
    contact_email = cfg.get("contact_email") or ""
    brand_name = cfg.get("brand_name") or "[TÊN THƯƠNG HIỆU]"
    signature = f"\nRegards,\n{contact_name}\n{contact_email}\n" if contact_name else "\nRegards,\n"
    detected_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rng = _draft_rng(domain)

    # URL bị phishing — dùng trong evidence section
    reported_url = target_url or f"https://{domain}"

    registrar = who.get("registrar") if isinstance(who, dict) else None
    abuse_emails = who.get("emails") if isinstance(who, dict) else None

    # T1: Lọc WHOIS privacy emails — chúng không xử lý report, gây lãng phí
    if isinstance(abuse_emails, list):
        filtered = [
            e for e in abuse_emails
            if e and not any(priv in e.lower() for priv in WHOIS_PRIVACY_DOMAINS)
        ]
        abuse_emails = ", ".join(filtered) if filtered else None
    elif isinstance(abuse_emails, str) and any(priv in abuse_emails.lower() for priv in WHOIS_PRIVACY_DOMAINS):
        abuse_emails = None

    # Fallback theo thứ tự ưu tiên: WHOIS emails → RDAP (ICANN standard) → static table
    abuse_email_source = "whois"
    if not abuse_emails:
        # Thử RDAP trước static table — RDAP là JSON có cấu trúc, đáng tin hơn
        rdap_result = get_rdap_abuse_email(domain)
        if rdap_result.get("abuse_email"):
            abuse_emails = rdap_result["abuse_email"]
            abuse_email_source = "rdap"
            # Nếu WHOIS không trả về registrar nhưng RDAP có → dùng tên từ RDAP
            if not registrar and rdap_result.get("registrar"):
                registrar = rdap_result["registrar"]
        elif registrar:
            fallback = lookup_registrar_abuse_email(registrar)
            if fallback:
                abuse_emails = fallback
                abuse_email_source = "static_table"

    # T3: Nếu registrar chỉ nhận web form, sinh file hướng dẫn thay vì email
    webform_url = None
    if registrar:
        r_lower = registrar.lower()
        for key, url in WEB_FORM_REGISTRARS.items():
            if key in r_lower:
                webform_url = url
                break

    if registrar and webform_url:
        path = os.path.join(REPORTS_DIR, f"{domain}_registrar_report.txt")
        contact_name = cfg.get("contact_name") or "[TÊN BẠN]"
        contact_email = cfg.get("contact_email") or "[EMAIL BẠN]"
        urlscan_link = f"https://urlscan.io/search/#page.domain:{domain}"
        vt_link_str = vt.get("link") or f"https://www.virustotal.com/gui/domain/{domain}"
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"""=== NỘI DUNG ĐỂ COPY VÀO WEB FORM: {webform_url} ===

--- FIELD: Domain / URL vi phạm ---
{reported_url}

--- FIELD: Loại vi phạm ---
Phishing / Brand Impersonation

--- FIELD: Mô tả (Description) ---
The domain {domain} is actively impersonating {brand_name} to deceive users into submitting sensitive information (login credentials, personal data, payment details).

Evidence:
- Phishing URL: {reported_url}
- VirusTotal report: {vt_link_str}
- URLScan analysis: {urlscan_link}
- First detected: {detected_date}

We request immediate suspension (serverHold / clientHold) of this domain.

--- FIELD: Thông tin liên hệ ---
Name: {contact_name}
Email: {contact_email}

--- FIELD: Bằng chứng đính kèm ---
[ĐÍNH KÈM ẢNH CHỤP MÀN HÌNH — bắt buộc có thanh địa chỉ trình duyệt]
(Ảnh chụp phải hiển thị rõ URL "{reported_url}" trên thanh địa chỉ của trình duyệt)
""")
        written.append(path)

    elif registrar:
        path = os.path.join(REPORTS_DIR, f"{domain}_registrar_report.txt")
        fallback_note = (
            "\n[NOTE: Abuse email looked up via RDAP (ICANN standard) — WHOIS did not return an email."
            " RDAP data is generally reliable, but verify before sending.]\n"
            if abuse_email_source == "rdap" else
            "\n[NOTE: Abuse email looked up from static registrar table — WHOIS and RDAP did not return an email."
            " Please verify this is correct before sending.]\n"
            if abuse_email_source == "static_table" else ""
        )
        subject = _pick(rng, _SUBJECT_REGISTRAR).format(domain=domain)
        opening = _pick(rng, _OPENING_REGISTRAR).format(domain=domain)
        description = _pick(rng, _DESCRIPTION_REGISTRAR)
        request = _pick(rng, _REQUEST_REGISTRAR)
        closing = _pick(rng, _CLOSING_REGISTRAR)

        # Evidence bullets — random thứ tự (trừ Domain luôn đứng đầu)
        vt_count = vt.get("malicious", 0) or 0
        vt_suspicious = vt.get("suspicious", 0) or 0
        vt_info = f"VirusTotal: {vt.get('link', 'N/A')}"
        if vt_count:
            vt_info += f" ({vt_count} security engines flagged as malicious)"
        elif vt_suspicious:
            vt_info += f" ({vt_suspicious} security engines flagged as suspicious)"

        extra_bullets = [
            f"Registrar: {registrar}",
            f"SSL Issuer: {cert.get('issuer', 'N/A')}",
            f"SSL Serial: {cert.get('serial', 'N/A')}",
            vt_info,
            f"First detected: {detected_date}",
        ]
        rng.shuffle(extra_bullets)
        evidence = "\n".join(f"- {b}" for b in [f"Domain: {domain}"] + extra_bullets)

        # Đoạn bổ sung chi tiết động — số engine VT + ngày phát hiện cụ thể
        _vt_sentences = []
        if vt_count:
            _vt_sentences = [
                f"As of {detected_date}, {vt_count} out of 91 security vendors on VirusTotal have flagged this domain as malicious.",
                f"This domain has been flagged by {vt_count} security engines on VirusTotal (as of {detected_date}), confirming active malicious activity.",
                f"VirusTotal analysis dated {detected_date} shows {vt_count} security vendors classifying this domain as a phishing threat.",
                f"Independent verification via VirusTotal ({detected_date}) confirms {vt_count} security vendors have identified this domain as malicious.",
            ]
        elif vt_suspicious:
            _vt_sentences = [
                f"As of {detected_date}, {vt_suspicious} security vendors on VirusTotal have flagged this domain as suspicious.",
                f"VirusTotal analysis dated {detected_date} shows {vt_suspicious} security vendors flagging this domain as suspicious.",
            ]
        vt_detail = _pick(rng, _vt_sentences) if _vt_sentences else ""

        _detection_sentences = [
            f"This domain was first identified by our security team on {detected_date} during routine brand monitoring.",
            f"Our security team detected this phishing site on {detected_date} through proactive brand protection monitoring.",
            f"We identified this threat on {detected_date} as part of our ongoing brand monitoring operations.",
            f"This domain came to our attention on {detected_date} during a routine brand impersonation scan.",
        ]
        detection_detail = _pick(rng, _detection_sentences)

        extra_context = "\n\n".join(filter(None, [vt_detail, detection_detail]))

        with open(path, "w", encoding="utf-8") as f:
            f.write(f"""To: {abuse_emails or '[TRA ABUSE EMAIL TẠI https://lookup.icann.org/]'}
Subject: {subject}
{fallback_note}
Dear {registrar} Abuse Team,

{opening}

{description}

{extra_context}

Evidence:
- Phishing URL  : {reported_url}
- Domain        : {domain}
{chr(10).join(f'- {b}' for b in extra_bullets)}

Screenshots: [ĐÍNH KÈM ẢNH CHỤP MÀN HÌNH — bắt buộc có thanh địa chỉ trình duyệt]
(Ảnh chụp phải hiển thị rõ URL "{reported_url}" trên thanh địa chỉ của trình duyệt)

{request}

We request serverHold / clientHold be placed on this domain to immediately stop the phishing operation.

{closing}
{signature}""")
        written.append(path)

    ca_note = match_ca_notes(cert.get("issuer"))
    if ca_note and ca_note.get("report_url"):
        path = os.path.join(REPORTS_DIR, f"{domain}_ca_report.txt")
        subject = _pick(rng, _SUBJECT_CA).format(domain=domain)
        opening = _pick(rng, _OPENING_CA).format(domain=domain)
        description = _pick(rng, _DESCRIPTION_CA)

        revoke_variants = [
            "Revoking this certificate will immediately degrade the phishing site's credibility and browser trust indicators. We kindly request expedited handling under your abuse/phishing revocation policy.",
            "Immediate revocation will remove the trust indicator that makes this phishing site appear legitimate to victims. Please process this under your emergency revocation policy.",
            "We urge expedited revocation to disrupt the phishing operation. Removing browser trust indicators will significantly reduce the impact on victims.",
            "Swift revocation will neutralize the SSL trust signal being exploited by this phishing campaign. Please prioritize under your abuse revocation procedures.",
        ]
        revoke_text = _pick(rng, revoke_variants)
        ca_detection = _pick(rng, [
            f"This phishing activity was first detected on {detected_date} during our brand monitoring operations.",
            f"Our security team identified this certificate being used for phishing on {detected_date}.",
            f"We discovered this fraudulent use of the certificate on {detected_date} via routine brand protection scanning.",
            f"This misuse of your issued certificate was detected on {detected_date} by our security team.",
        ])
        vt_ca_note = ""
        if vt_count:
            vt_ca_note = _pick(rng, [
                f"The domain has been confirmed malicious by {vt_count} security vendors on VirusTotal.",
                f"Independent VirusTotal analysis confirms {vt_count} security engines classify this domain as malicious.",
                f"{vt_count} out of 91 VirusTotal security vendors have flagged this domain as an active threat.",
            ])

        with open(path, "w", encoding="utf-8") as f:
            f.write(f"""Subject: {subject}

Dear {ca_note['ca'].title()} Security / Abuse Team,

{opening}

{description}

{ca_detection}{(" " + vt_ca_note) if vt_ca_note else ""}

Certificate details:
- Phishing URL   : {reported_url}
- Domain         : {domain}
- Serial Number  : {cert.get('serial', 'N/A')}
- Issued by      : {cert.get('issuer', 'N/A')}
- VirusTotal     : {vt.get('link', 'N/A')}
- First detected : {detected_date}

Screenshots: [ĐÍNH KÈM ẢNH CHỤP MÀN HÌNH — bắt buộc có thanh địa chỉ trình duyệt]
(Ảnh chụp phải hiển thị rõ URL "{reported_url}" trên thanh địa chỉ của trình duyệt)

{revoke_text}
{signature}""")
        written.append(path)

    return written


def export_nictop_excel(domains: list, brand_name: str = "", output_path: str = None) -> str:
    """Sinh file Excel theo template nic.top cho báo cáo phishing hàng loạt.

    nic.top nhận file Excel tối đa 200 domain/lần kèm ảnh chụp. Format chuẩn:
    - Cột A: Domain
    - Cột B: URL bị phishing
    - Cột C: Mô tả ngắn
    - Cột D: Brand bị giả mạo
    - Cột E: Ngày phát hiện

    Trả về đường dẫn file đã ghi.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError(
            "Cần cài openpyxl: pip install openpyxl"
        )

    os.makedirs(REPORTS_DIR, exist_ok=True)
    if output_path is None:
        output_path = os.path.join(REPORTS_DIR, f"nictop_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phishing Domains"

    # Header
    headers = ["Domain", "Phishing URL", "Description", "Brand Impersonated", "Date Detected"]
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    detected_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for row, domain in enumerate(domains[:200], 2):  # max 200
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=f"https://{domain}")
        ws.cell(row=row, column=3, value=f"Phishing site impersonating {brand_name or 'brand'} — harvesting user credentials")
        ws.cell(row=row, column=4, value=brand_name or "[BRAND NAME]")
        ws.cell(row=row, column=5, value=detected_date)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15

    wb.save(output_path)
    return output_path


def generate_apwg_draft(domain, vt, cfg):
    """Sinh sẵn email report gửi APWG eCrime (reportphishing@apwg.org).

    APWG (Anti-Phishing Working Group) là tổ chức phối hợp xử lý phishing quốc tế,
    nhận report qua email và phân phối tới các thành viên (ISP, hosting, browser vendors).
    Địa chỉ reportphishing@apwg.org đã được xác nhận hoạt động.
    """
    contact_name = cfg.get("contact_name") or ""
    contact_email = cfg.get("contact_email") or ""
    signature = f"\nRegards,\n{contact_name}\n{contact_email}\n" if contact_name else "\nRegards,\n"
    rng = _draft_rng(domain)

    _apwg_subjects = [
        "Phishing URL Report - {domain}",
        "Active Phishing Site Report - {domain}",
        "Brand Impersonation Phishing Report - {domain}",
        "Phishing Incident Report - {domain}",
    ]
    _apwg_openings = [
        "We are reporting the following URL as an active phishing site impersonating our brand to harvest user credentials and sensitive information.",
        "We have identified the following domain as an active phishing site targeting our brand and users.",
        "This is a phishing report for a site actively impersonating our brand and harvesting user credentials.",
        "We are submitting this report for an active phishing domain that is impersonating our brand's official website.",
    ]
    _apwg_descriptions = [
        "The site clones our brand's official login page to deceive users into submitting their credentials, OTP tokens, and payment details. Please add this URL to the APWG eCrime dataset to protect users across platforms.",
        "This phishing page replicates our brand's authentication portal to steal user login credentials and personal data. We request you add this to the eCrime dataset for cross-platform protection.",
        "The domain hosts a replica of our login interface designed to harvest credentials and OTP tokens from victims. Please include in the APWG eCrime dataset.",
        "Victims are tricked into submitting credentials, payment information, and personal data via this cloned login page. Adding this to the eCrime dataset will help protect users across member platforms.",
    ]

    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f"{domain}_apwg_report.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"""To: reportphishing@apwg.org
Subject: {_pick(rng, _apwg_subjects).format(domain=domain)}

Hello APWG eCrime Team,

{_pick(rng, _apwg_openings)}

Phishing URL: https://{domain}
Domain: {domain}
VirusTotal report: {vt.get('link', 'N/A')}

{_pick(rng, _apwg_descriptions)}
{signature}""")
    return path


def generate_hosting_draft(domain, ip, ip_whois, cfg):
    """Sinh draft DMCA/AUP takedown gửi ISP/hosting sở hữu IP gốc."""
    contact_name = cfg.get("contact_name") or ""
    contact_email = cfg.get("contact_email") or ""
    signature = f"\nRegards,\n{contact_name}\n{contact_email}\n" if contact_name else "\nRegards,\n"
    detected_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rng = _draft_rng(domain + "_hosting")

    os.makedirs(REPORTS_DIR, exist_ok=True)
    org = ip_whois.get("org") or "[TÊN TỔ CHỨC HOSTING CHƯA XÁC ĐỊNH]"
    abuse_email = ip_whois.get("abuse_email") or "[TRA ABUSE EMAIL TẠI https://lookup.icann.org/ hoặc rdap.arin.net]"
    asn = ip_whois.get("asn") or "N/A"

    subject = _pick(rng, [
        f"Phishing Site Takedown Request - {domain} (IP: {ip})",
        f"Abuse Report: Active Phishing Site on Your Network - {domain}",
        f"Hosting Abuse Report: Brand Impersonation Phishing - {domain}",
        f"Malicious Phishing Site Hosted on IP {ip} - {domain}",
    ])
    opening = _pick(rng, [
        f"We are writing to request the immediate suspension of a phishing website hosted on your infrastructure that is impersonating our brand.",
        f"We have identified an active phishing site hosted on your network (IP: {ip}) that is impersonating our brand and actively harvesting user credentials.",
        f"This is a formal takedown request for a phishing site running on your infrastructure (IP: {ip}) that fraudulently impersonates our brand.",
        f"We are reporting an active phishing operation hosted on your network that is cloning our brand's official website to steal user credentials.",
    ])
    description = _pick(rng, [
        f"The website at https://{domain} (hosted on IP {ip}) is an unauthorized clone of our brand's official platform, actively harvesting user credentials, OTP tokens, and payment details from unsuspecting visitors.",
        f"This phishing site at https://{domain} (IP: {ip}) replicates our brand's login interface to deceive users into submitting their credentials, OTP tokens, and financial information.",
        f"The domain {domain} (hosted at IP {ip}) hosts a fraudulent replica of our brand's authentication portal designed to steal login credentials and sensitive personal data from victims.",
        f"Hosted on IP {ip}, the site {domain} is a phishing page cloning our official brand portal to harvest user credentials, OTP codes, and payment details.",
    ])
    request = _pick(rng, [
        f"We request you suspend this hosting account immediately under your Acceptable Use Policy. This is causing ongoing harm to our users and brand reputation.",
        f"Please immediately suspend or null-route traffic to this IP/account under your AUP to prevent further credential theft from our users.",
        f"We urge you to terminate this hosting account without delay in accordance with your Acceptable Use Policy to stop the ongoing phishing operation.",
        f"Immediate suspension of this account/IP under your AUP is requested to protect users from ongoing credential harvesting.",
    ])

    path = os.path.join(REPORTS_DIR, f"{domain}_hosting_report.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"""To: {abuse_email}
Subject: {subject}

Dear {org} Abuse Team,

{opening}

{description}

Hosting details:
- Phishing domain: {domain}
- Origin IP: {ip}
- ASN: {asn}
- Hosting organization: {org}
- First detected: {detected_date}

Note: This IP was identified as a candidate origin server via subdomain
enumeration — please verify independently before taking action.

{request}
{signature}""")
    return path


def generate_registry_draft(domain, registry_info, cfg):
    """Sinh draft leo thang gửi Registry quản lý ccTLD."""
    if registry_info.get("source") == "not_found":
        return None

    # Xử lý trường hợp "không có kênh riêng" (.me, .tv, ...) — có trong bảng nhưng không có email/webform
    if (registry_info.get("source") == "static_table"
            and not registry_info.get("abuse_email")
            and not registry_info.get("report_webform")):
        return None  # Không sinh draft — guide bảo "báo nhà đăng ký"

    contact_name = cfg.get("contact_name") or ""
    contact_email = cfg.get("contact_email") or ""
    signature = f"\nRegards,\n{contact_name}\n{contact_email}\n" if contact_name else "\nRegards,\n"
    detected_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rng = _draft_rng(domain + "_registry")

    os.makedirs(REPORTS_DIR, exist_ok=True)
    if registry_info["source"] == "static_table":
        registry_name = registry_info.get("registry") or "Registry"
        abuse_email = registry_info.get("abuse_email") or "[TRA ABUSE EMAIL TẠI https://www.iana.org/domains/root/db]"
        note_line = f"\nSpecial note: {registry_info['note']}\n" if registry_info.get("note") else ""
        raw_section = ""
    else:
        registry_name = f"Registry (via {registry_info.get('whois_server', 'IANA')})"
        abuse_email = "[TRA ABUSE EMAIL TỪ NỘI DUNG WHOIS THÔ BÊN DƯỚI, hoặc https://www.iana.org/domains/root/db]"
        note_line = ""
        raw_section = (
            f"\n--- WHOIS raw from {registry_info.get('whois_server')} (for reference) ---\n"
            f"{registry_info.get('raw', '')}\n"
        )

    # T3: Nếu registry chỉ có web form (không có email), sinh file hướng dẫn
    webform_url_r = registry_info.get("report_webform")
    if webform_url_r and not registry_info.get("abuse_email"):
        registry_name_r = registry_info.get("registry") or "Registry"
        contact_name = cfg.get("contact_name") or "[TÊN BẠN]"
        contact_email = cfg.get("contact_email") or "[EMAIL BẠN]"
        vt_link_str = f"https://www.virustotal.com/gui/domain/{domain}"
        path = os.path.join(REPORTS_DIR, f"{domain}_registry_report.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"""=== NỘI DUNG ĐỂ COPY VÀO WEB FORM: {webform_url_r} ===

--- FIELD: Domain / URL vi phạm ---
{domain}

--- FIELD: Loại vi phạm ---
Phishing / Brand Impersonation

--- FIELD: Mô tả (Description) ---
The domain {domain} is actively used for phishing — impersonating a brand to deceive users into submitting credentials or personal data. We request immediate suspension (serverHold / clientHold).

Evidence:
- VirusTotal: {vt_link_str}
- First detected: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}
{note_line}
--- FIELD: Thông tin liên hệ ---
Name: {contact_name}
Email: {contact_email}

--- FIELD: Bằng chứng đính kèm ---
[ĐÍNH KÈM ẢNH CHỤP MÀN HÌNH — bắt buộc có thanh địa chỉ trình duyệt]
""")
        return path

    subject = _pick(rng, [
        f"Phishing Domain Suspension Request - {domain}",
        f"Registry Escalation: Active Phishing Domain - {domain}",
        f"Formal Complaint: Brand Impersonation Phishing Domain - {domain}",
        f"ClientHold Request for Phishing Domain - {domain}",
    ])
    opening = _pick(rng, [
        f"We are escalating a phishing domain abuse report directly to your registry regarding the domain {domain}, which is being used to impersonate our brand and harvest user credentials.",
        f"We are filing a formal escalation with your registry regarding {domain}, an active phishing domain impersonating our brand that has not been resolved at the registrar level.",
        f"This is a registry-level escalation for the phishing domain {domain}, which is actively impersonating our brand and stealing user credentials despite our reports to the registrar.",
        f"We are escalating directly to your registry to report {domain} as an active phishing domain that is fraudulently impersonating our brand.",
    ])
    escalation_reason = _pick(rng, [
        "We have already submitted abuse reports to the registrar, but are escalating to the registry level due to the ongoing harm and time-sensitive nature of this phishing campaign.",
        "Prior abuse reports to the domain registrar have not resulted in timely action, requiring this escalation to the registry level to protect our users.",
        "Given the active and ongoing nature of this phishing campaign and the urgency of protecting affected users, we are escalating directly to your registry.",
        "The continued operation of this phishing site necessitates registry-level intervention to ensure prompt suspension.",
    ])
    request = _pick(rng, [
        f"We respectfully request that {domain} be placed on ClientHold status immediately to stop active credential harvesting from users of our brand.",
        f"Please place {domain} on ClientHold (ICANN status) without delay to halt the ongoing phishing operation targeting our users.",
        f"We urge your registry to apply ClientHold to {domain} immediately under your abuse handling policy to stop the active phishing campaign.",
        f"Immediate ClientHold of {domain} is requested to neutralize this phishing threat and protect users from ongoing credential theft.",
    ])

    path = os.path.join(REPORTS_DIR, f"{domain}_registry_report.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"""To: {abuse_email}
Subject: {subject}

Dear {registry_name} Abuse Department,

{opening}

{escalation_reason}

{request}

Domain: {domain}
First detected: {detected_date}
{note_line}
This escalation is made in accordance with your registry's abuse handling
policy and ICANN compliance requirements.
{signature}{raw_section}""")
    return path


def generate_vncert_draft(domain, cert, vt, cfg):
    """Sinh draft report VNCERT — LUÔN sinh, người dùng tự quyết định có gửi không."""
    contact_name = cfg.get("contact_name") or ""
    contact_email = cfg.get("contact_email") or ""
    signature = f"\nTrân trọng,\n{contact_name}\n{contact_email}\n" if contact_name else "\nTrân trọng,\n"
    detected_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rng = _draft_rng(domain + "_vncert")
    vt_count = vt.get("malicious", 0) or 0

    subject = _pick(rng, [
        f"[Phản ánh phishing] {domain} giả mạo thương hiệu",
        f"[Báo cáo lừa đảo] Domain {domain} giả mạo nhãn hiệu",
        f"[Khẩn] Phản ánh website phishing {domain}",
        f"[Yêu cầu xử lý] {domain} — website giả mạo đang hoạt động",
    ])
    opening = _pick(rng, [
        f"Chúng tôi xin phản ánh domain {domain} đang thực hiện hành vi giả mạo thương hiệu của chúng tôi nhằm lừa đảo, đánh cắp thông tin nhạy cảm của người dùng (tài khoản đăng nhập, mã OTP, thông tin thẻ ngân hàng...).",
        f"Chúng tôi phát hiện domain {domain} đang giả mạo thương hiệu của chúng tôi để đánh cắp thông tin đăng nhập, mã OTP và dữ liệu tài chính của người dùng.",
        f"Chúng tôi báo cáo domain {domain} đang thực hiện tấn công phishing nhắm vào người dùng của chúng tôi bằng cách giả mạo giao diện chính thức của thương hiệu.",
        f"Chúng tôi kính báo VNCERT/CC về domain {domain} — một trang web phishing đang hoạt động, giả mạo thương hiệu chúng tôi nhằm đánh cắp thông tin người dùng.",
    ])
    description = _pick(rng, [
        "Domain giả mạo hoạt động bằng cách sao chép giao diện đăng nhập chính thức của thương hiệu chúng tôi, dụ người dùng nhập thông tin vào trang giả mạo.",
        "Trang web này sao chép giao diện đăng nhập của chúng tôi để lừa người dùng cung cấp tên đăng nhập, mật khẩu, mã OTP và thông tin thẻ ngân hàng.",
        "Website phishing này nhái lại thiết kế trang đăng nhập chính thống của chúng tôi, đánh lừa người dùng nhập thông tin nhạy cảm vào trang giả mạo.",
        "Kẻ tấn công đã tạo bản sao y hệt trang đăng nhập của chúng tôi tại domain này để thu thập thông tin xác thực và dữ liệu cá nhân của người dùng.",
    ])
    vt_note = ""
    if vt_count:
        vt_note = _pick(rng, [
            f"\nDomain này đã bị {vt_count} công cụ bảo mật trên VirusTotal gắn cờ là độc hại.",
            f"\nTheo VirusTotal, {vt_count} nhà cung cấp bảo mật đã xác nhận domain này là phishing/malicious.",
            f"\nKiểm tra trên VirusTotal ngày {detected_date} cho thấy {vt_count} engine bảo mật đã phát hiện domain này là mối đe dọa.",
        ])
    request = _pick(rng, [
        "Kính mong VNCERT/CC hỗ trợ xử lý, phối hợp với registrar/hosting để gỡ bỏ domain này nhằm bảo vệ người dùng Việt Nam.",
        "Chúng tôi kính đề nghị VNCERT/CC hỗ trợ can thiệp, yêu cầu registrar/hosting đình chỉ domain này để bảo vệ người dùng.",
        "Kính nhờ VNCERT/CC phối hợp xử lý khẩn cấp, liên hệ registrar/hosting để gỡ bỏ trang phishing này sớm nhất có thể.",
        "Chúng tôi mong nhận được sự hỗ trợ của VNCERT/CC trong việc yêu cầu đình chỉ domain phishing này nhằm ngăn chặn thiệt hại thêm cho người dùng.",
    ])

    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f"{domain}_vncert_report.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"""LƯU Ý: CHỈ gửi report này nếu domain nhắm vào nạn nhân tại Việt Nam
(vd. giả mạo thương hiệu/dịch vụ Việt Nam, nội dung tiếng Việt, nhắm vào người
dùng Việt Nam). Tool không tự xác định được điều này — tự đánh giá trước khi gửi.

To: report@vncert.vn
Subject: {subject}

Kính gửi VNCERT/CC,

{opening}

{description}{vt_note}

Thông tin kỹ thuật:
- Domain: {domain}
- SSL Issuer: {cert.get('issuer', 'N/A')}
- SSL Serial: {cert.get('serial', 'N/A')}
- VirusTotal: {vt.get('link', 'N/A')}
- Phát hiện lúc: {detected_date}

{request}
{signature}""")
    return path


# --------------------------------------------------------------------------
# Gửi báo cáo qua email thật (SMTP)
# --------------------------------------------------------------------------
#
# Chỉ gửi khi người dùng đã tự xác nhận thủ công (Streamlit: tick checkbox "đã đọc và xác
# nhận" trước khi nút "Gửi" hết disabled; CLI `send`: input("Gửi email này? (y/N): ")) — không
# có đường nào tự động gửi hàng loạt không qua duyệt, giữ đúng nguyên tắc xác minh-trước-khi-
# report xuyên suốt dự án (plan_phishing_takedown.md bước 2).


def _parse_proxy_url(proxy_str: str):
    """Parse chuỗi proxy thành tuple (proxy_type, host, port, username, password).

    Hỗ trợ: ``socks5://[user:pass@]host:port``, ``socks4://...``, ``http://...``,
    hoặc ``host:port`` (mặc định SOCKS5 khi không có scheme).
    Trả về None nếu PySocks chưa được cài.
    """
    if _socks is None:
        return None
    s = proxy_str.strip()
    proxy_type = _SOCKS5
    username = password = None

    if "://" in s:
        scheme, s = s.split("://", 1)
        scheme = scheme.lower()
        if scheme == "socks4":
            proxy_type = _SOCKS4
        elif scheme == "http":
            proxy_type = _SOCKS_HTTP
        # else: socks5 or unknown → SOCKS5

    if "@" in s:
        creds, s = s.rsplit("@", 1)
        username, _, password = creds.partition(":")

    host, _, port_str = s.rpartition(":")
    if not host:            # không có ":" → chỉ có host, dùng port mặc định
        host, port_str = s, "1080"
    return proxy_type, host.strip(), int(port_str.strip()), username or None, password or None


class _SMTPWithProxy(smtplib.SMTP):
    """smtplib.SMTP subclass routing traffic qua SOCKS/HTTP proxy (PySocks).

    Ghi đè _get_socket() — điểm kết nối TCP duy nhất mà smtplib.SMTP.connect() gọi —
    để tạo socks.socksocket() thay vì socket.create_connection() thông thường.
    self._proxy_info PHẢI được gán TRƯỚC super().__init__() vì __init__ gọi connect()
    ngay lập tức.
    """
    def __init__(self, host: str, port: int, proxy_info: tuple, timeout: float = 15):
        self._proxy_info = proxy_info   # ← gán trước, super().__init__ sẽ gọi _get_socket
        super().__init__(host, port, timeout=timeout)

    def _get_socket(self, host, port, timeout):
        proxy_type, p_host, p_port, p_user, p_pass = self._proxy_info
        sock = _socks.socksocket()
        sock.set_proxy(proxy_type, p_host, p_port, username=p_user, password=p_pass)
        sock.settimeout(timeout or socket.getdefaulttimeout())
        sock.connect((host, port))
        return sock


def _imap_save_sent(account: dict, raw_msg: bytes) -> str | None:
    """Lưu 1 bản copy email đã gửi vào thư mục Sent trên IMAP server.

    Chỉ chạy khi account có `imap_host` (hoặc dùng chung `host`), không raise —
    lỗi trả về chuỗi mô tả lỗi, thành công trả về None.
    Gmail tự lưu Sent khi gửi qua SMTP nên không cần — hàm này dành cho mail server
    tự host (vd. mail.camellrp.com) không tự động làm điều đó.
    """
    # Ưu tiên imap_host riêng, fallback dùng chung host với SMTP
    imap_host = account.get("imap_host") or account.get("host", "")
    imap_port = int(account.get("imap_port", 993))
    username = account.get("username", "")
    password = account.get("password", "")

    # Không làm gì với Gmail — Gmail tự lưu Sent qua SMTP
    if "gmail.com" in imap_host.lower():
        return None

    try:
        mail = imaplib.IMAP4_SSL(imap_host, imap_port)
        mail.login(username, password)

        # Tìm thư mục Sent — các mail server dùng tên khác nhau
        sent_folder = None
        _, folders = mail.list()
        for folder_line in (folders or []):
            name = folder_line.decode("utf-8", errors="ignore") if isinstance(folder_line, bytes) else str(folder_line)
            name_lower = name.lower()
            for candidate in ("sent", "sent items", "sent messages", "sent mail", "gesendete", "enviados"):
                if f'"{candidate}"' in name_lower or f"/{candidate}" in name_lower or name_lower.endswith(candidate):
                    # Trích tên folder thực từ chuỗi IMAP LIST response
                    parts = name.rsplit('"', 1)
                    sent_folder = parts[-1].strip().strip('"') if len(parts) > 1 else candidate
                    break
            if sent_folder:
                break
        if not sent_folder:
            sent_folder = "Sent"  # fallback mặc định

        # APPEND email vào Sent folder với flag \Seen
        now = imaplib.Time2Internaldate(datetime.now().timestamp())
        mail.append(sent_folder, r"\Seen", now, raw_msg)
        mail.logout()
        return None
    except Exception as e:
        return str(e)


def _send_via_account(account: dict, proxy_str: str | None, to: str, subject: str, body: str) -> dict:
    """Gửi email qua 1 SMTP account cụ thể, tùy chọn qua proxy.

    Tự detect mode:
    - port 465 hoặc ssl=true  → smtplib.SMTP_SSL (SSL/TLS ngay từ đầu, dùng cho mail.camellrp.com)
    - port 587 (mặc định)     → smtplib.SMTP + starttls() (STARTTLS, dùng cho Gmail)
    Không raise — trả về dict {"account", "proxy", "success", "error"}.
    """
    username = account.get("username", "")
    host = account.get("host", "smtp.gmail.com")
    port = int(account.get("port", 587))
    password = account.get("password", "")
    use_ssl = bool(account.get("ssl", False)) or port == 465
    proxy_label = proxy_str or "—"
    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = username
        msg["To"] = to

        to_list = [a.strip() for a in to.split(",") if a.strip()]

        if proxy_str:
            if _socks is None:
                raise RuntimeError("Proxy được cấu hình nhưng PySocks chưa cài (pip install PySocks)")
            proxy_info = _parse_proxy_url(proxy_str)
            if proxy_info is None:
                raise RuntimeError(f"Không parse được proxy: {proxy_str}")
            server_ctx = _SMTPWithProxy(host, port, proxy_info, timeout=15)
            with server_ctx as server:
                if not use_ssl:
                    server.starttls()
                server.login(username, password)
                server.sendmail(username, to_list, msg.as_string())
        elif use_ssl:
            with smtplib.SMTP_SSL(host, port, timeout=15) as server:
                server.login(username, password)
                server.sendmail(username, to_list, msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=15) as server:
                server.starttls()
                server.login(username, password)
                server.sendmail(username, to_list, msg.as_string())

        # Lưu copy vào Sent folder qua IMAP (không làm gì với Gmail — tự lưu)
        imap_err = _imap_save_sent(account, msg.as_bytes())
        imap_note = f" (IMAP Sent: {imap_err})" if imap_err else ""

        return {"account": username, "proxy": proxy_label, "success": True, "error": None, "imap_note": imap_note}
    except Exception as e:
        return {"account": username, "proxy": proxy_label, "success": False, "error": str(e)}


def send_report_email_bulk(to: str, subject: str, body: str, cfg: dict) -> list:
    """Gửi email qua TẤT CẢ smtp_accounts đã cấu hình, mỗi account dùng 1 proxy xoay vòng
    (cycle nếu ít proxy hơn account), tất cả chạy ĐỒNG THỜI (ThreadPoolExecutor).

    Trả về list dict — 1 phần tử / account: {"account", "proxy", "success", "error"}.
    Không raise — nơi gọi luôn nhận được list để render kết quả / ghi log.
    """
    accounts = cfg.get("smtp_accounts", [])
    proxies = cfg.get("smtp_proxies", [])

    if not accounts:
        return [{"account": "—", "proxy": "—", "success": False,
                 "error": "Chưa cấu hình smtp_accounts trong config.ini"}]

    # Pair mỗi account với proxy xoay vòng (vd. 3 account + 2 proxy → proxy 0, 1, 0)
    pairs = [
        (account, proxies[i % len(proxies)] if proxies else None)
        for i, account in enumerate(accounts)
    ]

    with concurrent.futures.ThreadPoolExecutor(max_workers=len(pairs)) as pool:
        futures = {
            pool.submit(_send_via_account, account, proxy, to, subject, body): i
            for i, (account, proxy) in enumerate(pairs)
        }
        # Giữ thứ tự theo index account (as_completed trả về không theo thứ tự)
        indexed = {}
        for future in concurrent.futures.as_completed(futures):
            indexed[futures[future]] = future.result()
    return [indexed[i] for i in sorted(indexed)]


def send_report_email_single(to: str, subject: str, body: str, account: dict, proxy_str: str | None = None) -> dict:
    """Gửi email qua 1 account chỉ định (dùng khi UI cho phép chọn account cụ thể — P3).

    Trả về dict {"account", "proxy", "success", "error"} — không raise.
    """
    return _send_via_account(account, proxy_str, to, subject, body)

# file khi ghi sent_log.csv, không liên quan gì tới parse_draft_email() bên dưới.
_DRAFT_FILENAME_SUFFIXES = [
    "_registrar_report.txt",
    "_ca_report.txt",
    "_apwg_report.txt",
    "_hosting_report.txt",
    "_registry_report.txt",
    "_vncert_report.txt",
]


def domain_from_draft_filename(filename: str) -> str:
    for suffix in _DRAFT_FILENAME_SUFFIXES:
        if filename.endswith(suffix):
            return filename[: -len(suffix)]
    return os.path.splitext(filename)[0]


def parse_draft_email(path: str) -> dict:
    """Đọc 1 file draft đã sinh bởi generate_*_draft() ở trên, tách ra {"to", "subject", "body"}.

    KHÔNG giả định "To:"/"Subject:" luôn nằm ở 2 dòng đầu file: generate_vncert_draft() ghi 1
    đoạn cảnh báo ("LƯU Ý: CHỈ gửi report này nếu...") đứng TRƯỚC khối header, có dòng trống ở
    giữa — nếu chỉ lấy "dòng trống đầu tiên trong cả file" làm ranh giới body thì sẽ cắt nhầm
    ngay sau đoạn cảnh báo, làm "to"/"subject" thật ở phía dưới lẫn vào phần đầu của "body". Vì
    vậy hàm này quét toàn file tìm dòng "To:" đầu tiên (có thể không tồn tại — draft CA report
    của generate_email_drafts() cố ý không có dòng này vì nhiều CA dùng web form thay vì email)
    và dòng "Subject:" đầu tiên (luôn tồn tại, trừ file hỏng), rồi lấy body là phần sau dòng
    trống đầu tiên NGAY SAU dòng Subject đó — không phải dòng trống đầu tiên của cả file.

    "to" trả về None (coi như KHÔNG gửi được qua email) nếu dòng "To:" không tồn tại, hoặc giá
    trị của nó là 1 placeholder cần tra cứu thủ công (chứa "[TRA ABUSE EMAIL" hoặc "[KHÔNG CÓ")
    — không được cố gửi tới 1 chuỗi placeholder như vậy.
    """
    with open(path, encoding="utf-8") as f:
        lines = f.read().split("\n")

    to = None
    subject = None
    subject_idx = None
    for i, line in enumerate(lines):
        if to is None and line.startswith("To:"):
            to = line[len("To:"):].strip()
        if subject is None and line.startswith("Subject:"):
            subject = line[len("Subject:"):].strip()
            subject_idx = i
            break  # "Subject:" luôn đứng sau "To:" (nếu có) trong mọi draft hiện tại — dừng quét

    if subject_idx is None:
        return {"to": None, "subject": None, "body": "\n".join(lines)}

    body_start = subject_idx + 1
    for i in range(subject_idx + 1, len(lines)):
        if lines[i].strip() == "":
            body_start = i + 1
            break
    body = "\n".join(lines[body_start:])

    if to and ("[TRA ABUSE EMAIL" in to or "[KHÔNG CÓ" in to):
        to = None
    if to:
        to = _normalize_to_addresses(to)

    return {"to": to, "subject": subject, "body": body}


def _normalize_to_addresses(to: str) -> str:
    """Chuẩn hóa trường To: về dạng 'addr1, addr2' thuần túy.

    Xử lý các trường hợp:
    - Python list repr: "['addr1', 'addr2']" — do python-whois trả về list, bị ghi nguyên vào file
    - Slash-separated: "addr1 / addr2" — legacy format trong CCTLD_REGISTRY_CONTACTS (vd. EURid, RU)
    - String đơn: giữ nguyên

    Trả về chuỗi comma-separated sạch để dùng trong header email và sendmail().
    """
    import ast
    if to and to.startswith("["):
        try:
            parsed = ast.literal_eval(to)
            if isinstance(parsed, list):
                return ", ".join(str(e).strip() for e in parsed if e)
        except Exception:
            pass
    if to and " / " in to:
        parts = [p.strip() for p in to.split(" / ") if p.strip()]
        return ", ".join(parts)
    return to
    """Gửi 1 email report thật qua SMTP (smtplib, thư viện chuẩn, không cần cài thêm gì).

    Chỉ được gọi SAU KHI người dùng đã xác nhận thủ công ở nơi gọi (xem comment đầu section
    này) — hàm này tự nó không có gate xác nhận nào, không dùng trực tiếp từ nơi khác mà bỏ
    qua bước xác nhận đó. Bọc try/except toàn bộ và không raise ra ngoài, để nơi gọi luôn nhận
    được dict để hiển thị/ghi log thay vì làm crash Streamlit hoặc CLI.

    Backward compat: dùng account đầu tiên trong smtp_accounts (nếu có), hoặc fallback về
    smtp_host/username/password cũ. Gửi qua proxy đầu tiên nếu smtp_proxies được cấu hình.
    """
    accounts = cfg.get("smtp_accounts", [])
    proxies = cfg.get("smtp_proxies", [])
    if accounts:
        result = _send_via_account(accounts[0], proxies[0] if proxies else None, to, subject, body)
        return {"success": result["success"], "error": result.get("error")}
    # Fallback về keys cũ
    if not cfg.get("smtp_host") or not cfg.get("smtp_username") or not cfg.get("smtp_password"):
        return {
            "success": False,
            "error": "Chưa cấu hình đủ [smtp] trong config.ini (host/username/password)",
        }
    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = cfg["smtp_username"]
        msg["To"] = to
        to_list = [a.strip() for a in to.split(",") if a.strip()]

        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"], timeout=15) as server:
            server.starttls()
            server.login(cfg["smtp_username"], cfg["smtp_password"])
            server.sendmail(cfg["smtp_username"], to_list, msg.as_string())
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


@contextmanager
def sent_log_lock(timeout: float = 15.0, stale_after: float = 120.0):
    """Cross-process lock protecting a shared sent_log.csv and its schema."""
    lock_path = SENT_LOG_PATH + ".lock"
    deadline = time.monotonic() + timeout
    while True:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            try:
                os.write(fd, f"{os.getpid()}\n{time.time()}\n".encode("ascii"))
            finally:
                os.close(fd)
            break
        except FileExistsError:
            try:
                if time.time() - os.path.getmtime(lock_path) > stale_after:
                    os.remove(lock_path)
                    continue
            except FileNotFoundError:
                continue
            if time.monotonic() >= deadline:
                raise TimeoutError(f"Không thể khóa file log dùng chung: {lock_path}")
            time.sleep(0.05)
    try:
        yield
    finally:
        try:
            os.remove(lock_path)
        except FileNotFoundError:
            pass


def log_sent(row: dict):
    """Ghi 1 dòng vào sent_log.csv (file RIÊNG, không phải cột mới trong case_log.csv — giữ
    đúng nguyên tắc không đổi schema CSV đã có dữ liệu thật, đã áp dụng cho reputation/
    cdn_detected/origin_ip_scan/registry_contact trước đó). Gọi cho mọi lần gửi, kể cả thất
    bại, để sent_log.csv là nhật ký đầy đủ của mọi lần ai đó đã bấm gửi.
    """
    with sent_log_lock():
        is_new = not os.path.exists(SENT_LOG_PATH)
        if not is_new:
            with open(SENT_LOG_PATH, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                existing_fields = reader.fieldnames or []
                existing_rows = list(reader)
            missing_fields = [key for key in row if key not in existing_fields]
            if missing_fields:
                fields = existing_fields + missing_fields
                temp_path = f"{SENT_LOG_PATH}.{os.getpid()}.tmp"
                try:
                    with open(temp_path, "w", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=fields)
                        writer.writeheader()
                        writer.writerows(existing_rows)
                        f.flush()
                        os.fsync(f.fileno())
                    os.replace(temp_path, SENT_LOG_PATH)
                finally:
                    try:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    except OSError:
                        pass
        with open(SENT_LOG_PATH, "a", newline="", encoding="utf-8") as f:
            if is_new:
                fields = list(row.keys())
            else:
                with open(SENT_LOG_PATH, newline="", encoding="utf-8-sig") as existing:
                    fields = next(csv.reader(existing), list(row.keys()))
            writer = csv.DictWriter(f, fieldnames=fields)
            if is_new:
                writer.writeheader()
            writer.writerow(row)
            f.flush()
            os.fsync(f.fileno())


def append_reported_url_to_drafts(drafts: list, target_url: str) -> list:
    """Ensure every generated report contains the exact offending URL/path."""
    marker = f"Reported URL: {target_url}"
    updated = []
    for path in drafts:
        try:
            with open(path, encoding="utf-8") as f:
                content = f.read()
            if marker not in content:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(content.rstrip() + f"\n\n{marker}\n")
            updated.append(path)
        except OSError:
            updated.append(path)
    return updated


# --------------------------------------------------------------------------
# Commands
# --------------------------------------------------------------------------

def run_cdn_check(target: str) -> dict:
    """Pipeline tối giản: phát hiện Cloudflare/CDN và registrar, không gọi API key nào."""
    domain = normalize_domain(target)
    who = get_whois_info(domain)
    cf = is_cloudflare(who.get("name_servers")) if isinstance(who, dict) else False
    registrar = who.get("registrar") if isinstance(who, dict) else None
    try:
        cdn_detected = detect_cdn(domain)
    except Exception:
        cdn_detected = []
    # Tra TLD registry từ bảng tĩnh (không cần mạng, không fallback IANA)
    registry_contact = _static_registry_lookup(domain)
    return {
        "domain": domain,
        "cloudflare": cf,
        "cdn_detected": cdn_detected,
        "registrar": registrar,
        "registry_contact": registry_contact,
    }


def _static_registry_lookup(domain: str) -> dict | None:
    """Tra CCTLD_REGISTRY_CONTACTS từ bảng tĩnh, thử 2-level trước rồi 1-level.
    Trả về dict entry hoặc None nếu không tìm thấy."""
    d = domain.lower()
    parts = d.rsplit(".", 2)
    if len(parts) >= 3:
        two_level = f"{parts[-2]}.{parts[-1]}"
        if two_level in CCTLD_REGISTRY_CONTACTS:
            return CCTLD_REGISTRY_CONTACTS[two_level]
    tld = d.rsplit(".", 1)[-1]
    return CCTLD_REGISTRY_CONTACTS.get(tld)


def playwright_available() -> bool:
    """Kiểm tra Playwright đã cài chưa (không import nặng nếu chưa có)."""
    import importlib.util
    return importlib.util.find_spec("playwright") is not None


_REPORT_BROWSER_MANAGER = None
_REPORT_BROWSER_MANAGER_LOCK = None


def _target_url(value: str) -> str:
    """Giữ nguyên URL/path; chỉ thêm https:// khi đầu vào là domain thuần."""
    value = str(value or "").strip()
    return value if "://" in value else f"https://{normalize_domain(value)}"


class _ReportBrowserManager:
    """Một Playwright browser dùng chung; mỗi report được mở thành tab mới.

    Browser được relaunch khi dark_mode thay đổi vì --force-dark-mode là flag
    browser-level, không thể đổi sau khi đã launch.
    """

    def __init__(self):
        import queue
        import threading

        self._queue = queue.Queue()
        self._thread = threading.Thread(target=self._run, daemon=True, name="report-browser")
        self._thread.start()

    def open(self, kind: str, dark_mode: bool = True, **payload) -> dict:
        import threading

        done = threading.Event()
        request = {"kind": kind, "dark_mode": dark_mode, "payload": payload, "done": done, "result": None}
        self._queue.put(request)
        if not done.wait(timeout=40):
            return {"error": "Chrome mở quá lâu. Hãy thử lại."}
        return request["result"]

    def _run(self):
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            while True:
                request = self._queue.get()
                request["result"] = {
                    "error": "playwright chưa cài — chạy: pip install playwright && python -m playwright install chromium"
                }
                request["done"].set()

        pw = sync_playwright().start()
        browser = None
        current_dark_mode = None  # track theme browser đang dùng

        while True:
            request = self._queue.get()
            dark_mode = request.get("dark_mode", True)
            try:
                # Relaunch nếu theme thay đổi hoặc browser bị đóng
                if browser is None or not browser.is_connected() or current_dark_mode != dark_mode:
                    if browser is not None:
                        try:
                            browser.close()
                        except Exception:
                            pass
                    args = ["--start-maximized"]
                    if dark_mode:
                        # --force-dark-mode: Chrome UI tối (toolbar, titlebar)
                        # WebContentsForceDark: force dark cho trang không support media query
                        args += ["--force-dark-mode", "--enable-features=WebContentsForceDark"]
                    browser = pw.chromium.launch(headless=False, slow_mo=250, args=args)
                    current_dark_mode = dark_mode

                color_scheme = "dark" if dark_mode else "light"
                page = browser.new_page(color_scheme=color_scheme, no_viewport=True)
                if request["kind"] == "google":
                    self._fill_google(page, **request["payload"])
                elif request["kind"] == "microsoft":
                    self._fill_microsoft(page, **request["payload"])
                else:
                    raise ValueError(f"Loại form không hỗ trợ: {request['kind']}")
                page.bring_to_front()
                request["result"] = {"status": "launched"}
            except Exception as exc:
                request["result"] = {"error": str(exc)}
            finally:
                request["done"].set()

    @staticmethod
    def _fill_google(page, target_url: str, description: str,
                     threat_type: str, threat_category: str):
        import urllib.parse as _up

        gsb_url = (
            "https://safebrowsing.google.com/safebrowsing/report_phish/?url="
            + _up.quote(target_url, safe="")
        )
        page.goto(gsb_url, wait_until="domcontentloaded", timeout=20_000)

        url_field = page.locator('[formcontrolname="url"]')
        url_field.wait_for(state="visible", timeout=10_000)
        url_field.fill(target_url)

        details = page.locator('[formcontrolname="details"]')
        details.wait_for(state="visible", timeout=10_000)
        details.fill(description)

        l1 = page.locator('[formcontrolname="l1Taxonomy"]')
        l1.click()
        page.get_by_role("option", name=threat_type, exact=True).click()

        if threat_category and threat_category != "None":
            l3 = page.locator('[formcontrolname="l3Taxonomy"]')
            l3.wait_for(state="visible", timeout=5_000)
            l3.click()
            page.get_by_role("option", name=threat_category, exact=True).click()

    @staticmethod
    @staticmethod
    def _fill_microsoft(page, target_url: str):
        page.goto(
            "https://www.microsoft.com/en-us/wdsi/support/report-unsafe-site-guest",
            wait_until="domcontentloaded",
            timeout=20_000,
        )
        url_field = page.locator("#WebsiteUrlOrIP")
        url_field.wait_for(state="visible", timeout=10_000)
        url_field.fill(target_url)

        page.locator("#LanguageListButton").click()
        vietnamese = page.locator('[role="option"][aria-label="Vietnamese"]')
        vietnamese.wait_for(state="visible", timeout=5_000)
        vietnamese.click()


def _report_browser_manager() -> _ReportBrowserManager:
    global _REPORT_BROWSER_MANAGER, _REPORT_BROWSER_MANAGER_LOCK
    if _REPORT_BROWSER_MANAGER_LOCK is None:
        import threading
        _REPORT_BROWSER_MANAGER_LOCK = threading.Lock()
    with _REPORT_BROWSER_MANAGER_LOCK:
        if _REPORT_BROWSER_MANAGER is None:
            _REPORT_BROWSER_MANAGER = _ReportBrowserManager()
    return _REPORT_BROWSER_MANAGER


def open_gsb_form_playwright(target: str, description: str,
                             threat_type: str = "Social Engineering",
                             threat_category: str = "None",
                             dark_mode: bool = True) -> dict:
    """Mở tab Google trong browser automation dùng chung và tự điền form."""
    return _report_browser_manager().open(
        "google",
        dark_mode=dark_mode,
        target_url=_target_url(target),
        description=description,
        threat_type=threat_type,
        threat_category=threat_category,
    )


def report_netcraft_api(target_url: str, reason: str, email: str) -> dict:
    """Gửi report URL lên Netcraft qua API v3 (không cần API key, không cần Playwright).

    https://report.netcraft.com/api/v3#tag/Report/operation/report_urls
    Trả về dict gồm payload gửi đi, status_code, response body để kiểm tra.
    """
    if requests is None:
        return {"error": "requests not installed"}
    payload = {
        "email": email,
        "urls": [{
            "url": target_url,
            "country": "VN",
            "reason": reason,
        }],
    }
    try:
        r = requests.post(
            "https://report.netcraft.com/api/nightly/report/urls",
            json=payload,
            timeout=15,
        )
        try:
            resp_body = r.json()
        except Exception:
            resp_body = r.text
        result = {
            "payload": payload,
            "status_code": r.status_code,
            "response": resp_body,
        }
        if r.status_code in (200, 201, 204):
            result["success"] = True
        else:
            result["error"] = f"HTTP {r.status_code}"
        return result
    except Exception as exc:
        return {"payload": payload, "error": str(exc)}


def open_microsoft_form_playwright(target: str, dark_mode: bool = True) -> dict:
    """Mở tab Microsoft trong browser automation dùng chung và tự điền form."""
    return _report_browser_manager().open("microsoft", dark_mode=dark_mode, target_url=_target_url(target))


def open_cloudflare_form_browser(domain: str) -> dict:
    """Mở form Cloudflare Abuse trong browser thật của user.

    abuse.cloudflare.com dùng chính bot-detection của Cloudflare nên Playwright
    bị block ngay lập tức. Giải pháp duy nhất đáng tin cậy là mở bằng browser
    thật — user tự paste description từ nút Copy bên cạnh.
    """
    import webbrowser
    url = "https://abuse.cloudflare.com/phishing"
    webbrowser.open(url)
    return {"status": "launched"}



def run_check(target: str, submit: bool, cfg: dict) -> dict:
    """Chạy toàn bộ pipeline kiểm tra 1 domain và trả về dict kết quả đầy đủ.

    NGUỒN LOGIC DUY NHẤT cho lệnh `check`: cả CLI (cmd_check) và Streamlit UI
    gọi thẳng hàm này rồi tự quyết định cách hiển thị, để không bao giờ lệch
    kết quả giữa hai giao diện.
    """
    import concurrent.futures as _cf
    domain = normalize_domain(target)
    target_url = target.strip() if "://" in target else f"http://{domain}"

    # URLScan: submit sớm (song song với pipeline chính) để thời gian ~30s xử lý
    # trùng với các bước WHOIS/VT/GSB bên dưới — join ở cuối khi gần xong rồi.
    _urlscan_key = cfg.get("urlscan_api_key", "")
    _urlscan_executor = _cf.ThreadPoolExecutor(max_workers=1)
    if _urlscan_key:
        _urlscan_future = _urlscan_executor.submit(urlscan_submit_and_wait, domain, _urlscan_key)
    else:
        _urlscan_future = None

    try:
        cert = get_cert_info(domain)
    except Exception as e:
        cert = {"error": str(e)}

    who = get_whois_info(domain)
    cf = is_cloudflare(who.get("name_servers")) if isinstance(who, dict) else False
    domain_age_days = compute_domain_age_days(who)

    try:
        http_check = check_http(target_url)
    except Exception as e:
        http_check = {"error": str(e)}

    try:
        mx_records = check_mx_records(domain)
    except Exception as e:
        mx_records = {"records": [], "providers": [], "error": str(e)}

    # detect_cdn()/scan_common_subdomains() tự bọc try/except nội bộ (per-nguồn tín hiệu,
    # per-subdomain) nên hiếm khi raise ra ngoài, nhưng vẫn bọc thêm 1 lớp ở đây để lỗi bất ngờ
    # (vd. domain không resolve được gì cả) không làm hỏng kết quả SSL/WHOIS/VirusTotal đã lấy
    # được ở trên — cùng nguyên tắc với log_case/generate_email_drafts bên dưới.
    try:
        cdn_detected = detect_cdn(domain)
    except Exception:
        cdn_detected = []

    try:
        origin_ip_scan = scan_common_subdomains(domain)
    except Exception:
        origin_ip_scan = {}

    vt = check_virustotal(domain, cfg["vt_api_key"])
    virustotal_submit = None
    if vt.get("note") == "Domain chưa từng được VirusTotal quét trước đây" and submit:
        virustotal_submit = submit_virustotal(domain, cfg["vt_api_key"])

    gsb = check_safebrowsing(target_url, cfg["gsb_api_key"])

    reputation = compute_reputation(vt, gsb)

    ca_note = match_ca_notes(cert.get("issuer"))

    log_row = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "domain": domain,
        "target_url": target_url,
        "ip": cert.get("ip"),
        "ssl_issuer": cert.get("issuer"),
        "ssl_serial": cert.get("serial"),
        "registrar": who.get("registrar") if isinstance(who, dict) else None,
        "cloudflare": cf,
        "vt_malicious": vt.get("malicious"),
        "vt_link": vt.get("link"),
        "gsb_flagged": gsb.get("flagged"),
        "status": "detected",
    }
    # log_case/generate_email_drafts chỉ ghi file phụ trợ (log, email draft) — nếu
    # chúng lỗi (vd. case_log.csv đang bị khóa bởi Excel), phần điều tra
    # SSL/WHOIS/VirusTotal/Safe Browsing đã lấy được ở trên vẫn phải trả về, không
    # được để exception ở bước ghi file làm mất toàn bộ kết quả đã có.
    try:
        log_case(log_row)
        log_error = None
    except Exception as e:
        log_error = str(e)

    try:
        drafts = generate_email_drafts(domain, cert, who, vt, cfg, target_url=target_url)
        drafts_error = None
        # Tính toán nguồn email registrar để UI có thể hiển thị (whois/rdap/static_table)
        _who_emails = who.get("emails") if isinstance(who, dict) else None
        if isinstance(_who_emails, list):
            _who_emails_filtered = [
                e for e in _who_emails
                if e and not any(priv in e.lower() for priv in WHOIS_PRIVACY_DOMAINS)
            ]
            _who_emails = ", ".join(_who_emails_filtered) if _who_emails_filtered else None
        _registrar = who.get("registrar") if isinstance(who, dict) else None
        if _who_emails:
            registrar_abuse_email_source = "whois"
            registrar_abuse_email_used = _who_emails
        else:
            rdap_r = get_rdap_abuse_email(domain)
            if rdap_r.get("abuse_email"):
                registrar_abuse_email_source = "rdap"
                registrar_abuse_email_used = rdap_r["abuse_email"]
            elif _registrar and lookup_registrar_abuse_email(_registrar):
                registrar_abuse_email_source = "static_table"
                registrar_abuse_email_used = lookup_registrar_abuse_email(_registrar)
            else:
                registrar_abuse_email_source = None
                registrar_abuse_email_used = None
    except Exception as e:
        drafts = []
        drafts_error = str(e)
        registrar_abuse_email_source = None
        registrar_abuse_email_used = None

    # T1: APWG (reportphishing@apwg.org) chỉ là kho dữ liệu — không có quyền gỡ domain.
    # Đã gửi 743 emails → 0 kết quả (dữ liệu thực tế 35 ngày). Bỏ khỏi pipeline.
    # Hàm generate_apwg_draft() vẫn giữ cho người dùng gọi thủ công nếu muốn.
    # try:
    #     apwg_draft = generate_apwg_draft(domain, vt, cfg)
    #     drafts.append(apwg_draft)
    # except Exception as e:
    #     drafts_error = (drafts_error + "; " if drafts_error else "") + f"APWG draft: {e}"

    # Hosting/ISP takedown (03_Technical_Guide.md mục 3) — CHỈ chạy khi origin_ip_scan (ở trên)
    # tìm thấy ít nhất 1 subdomain có IP KHÁC cert["ip"]. Đặt SAU khối openphish (không phải ngay
    # sau origin_ip_scan) vì cần drafts/drafts_error đã tồn tại để append/gộp lỗi vào — đặt sớm
    # hơn sẽ NameError. Nếu có nhiều candidate IP, chỉ xử lý 1 IP đầu tiên (sort để ổn định giữa
    # các lần chạy) để tránh spam nhiều file draft cho 1 domain — xem CLAUDE.md.
    origin_ip_whois = {}
    candidate_ips = {ip for ip in origin_ip_scan.values() if ip and ip != cert.get("ip")}
    if candidate_ips:
        try:
            first_ip = sorted(candidate_ips)[0]
            origin_ip_whois = {first_ip: get_ip_whois(first_ip)}
            hosting_draft = generate_hosting_draft(domain, first_ip, origin_ip_whois[first_ip], cfg)
            drafts.append(hosting_draft)
        except Exception as e:
            drafts_error = (drafts_error + "; " if drafts_error else "") + f"Hosting draft: {e}"

    # Leo thang Registry (ccTLD lạ, 03_Technical_Guide.md mục 5) + report VNCERT (mục 8). Cùng
    # pattern try/except với các khối draft ở trên — lỗi ở đây không được phá hỏng kết quả điều
    # tra đã có. registry_info luôn được trả về (kể cả "not_found"), vncert_draft luôn được sinh
    # (không điều kiện gì, giống openphish) vì tool không tự biết domain có nhắm vào nạn nhân VN.
    try:
        registry_info = lookup_registry_contact(domain)
        registry_draft = generate_registry_draft(domain, registry_info, cfg)
        if registry_draft:
            drafts.append(registry_draft)
        vncert_draft = generate_vncert_draft(domain, cert, vt, cfg)
        drafts.append(vncert_draft)
    except Exception as e:
        registry_info = {"source": "not_found"}
        drafts_error = (drafts_error + "; " if drafts_error else "") + f"Registry/VNCERT draft: {e}"

    drafts = append_reported_url_to_drafts(drafts, target_url)

    # Thu kết quả URLScan (thread đã chạy song song từ đầu — thường đã xong hoặc sắp xong)
    urlscan_data = {}
    if _urlscan_future is not None:
        try:
            urlscan_data = _urlscan_future.result(timeout=75)  # đã poll nội bộ rồi, chỉ cần thêm buffer nhỏ
        except Exception as e:
            urlscan_data = {"error": str(e)}
        finally:
            _urlscan_executor.shutdown(wait=False)
        # Gắn evidence vào tất cả draft nếu URLScan thành công (hoặc ít nhất có screenshot_url)
        if urlscan_data.get("screenshot_url"):
            append_urlscan_evidence_to_drafts(drafts, urlscan_data)
    else:
        _urlscan_executor.shutdown(wait=False)

    return {
        "domain": domain,
        "cert": cert,
        "whois": who,
        "cloudflare": cf,
        "cdn_detected": cdn_detected,
        "origin_ip_scan": origin_ip_scan,
        "origin_ip_whois": origin_ip_whois,
        "registry_contact": registry_info,
        "virustotal": vt,
        "virustotal_submit": virustotal_submit,
        "safebrowsing": gsb,
        "reputation": reputation,
        "ca_note": ca_note,
        "http_check": http_check,
        "domain_age_days": domain_age_days,
        "mx_records": mx_records,
        "log_row": log_row,
        "log_error": log_error,
        "drafts": drafts,
        "drafts_error": drafts_error,
        "registrar_abuse_email_source": registrar_abuse_email_source,
        "registrar_abuse_email_used": registrar_abuse_email_used,
        "urlscan": urlscan_data,
    }


def cmd_check(args):
    cfg = load_config()
    result = run_check(args.target, args.submit, cfg)
    domain = result["domain"]
    cert = result["cert"]
    who = result["whois"]
    cf = result["cloudflare"]
    cdn_detected = result["cdn_detected"]
    origin_ip_scan = result["origin_ip_scan"]
    vt = result["virustotal"]
    gsb = result["safebrowsing"]
    ca_note = result["ca_note"]

    print(f"\n=== Đang kiểm tra: {domain} ===\n")

    print("--- SSL Certificate ---")
    if "error" in cert:
        print(f"  Lỗi: {cert['error']}")
    else:
        for k, v in cert.items():
            print(f"  {k:15s}: {v}")

    print("\n--- WHOIS ---")
    for k, v in who.items():
        print(f"  {k:15s}: {v}")

    print(f"\n--- Cloudflare ---\n  {'Có' if cf else 'Không rõ / Không'}")

    print("\n--- VirusTotal ---")
    for k, v in vt.items():
        print(f"  {k:15s}: {v}")
    if result["virustotal_submit"]:
        print("  " + result["virustotal_submit"])

    print("\n--- Google Safe Browsing ---")
    for k, v in gsb.items():
        print(f"  {k:15s}: {v}")

    rep = result["reputation"]
    print(f"\n--- Uy tín (tổng hợp VirusTotal + Safe Browsing) ---\n  {rep['label']}")
    for reason in rep["reasons"]:
        print(f"    - {reason}")
    if rep["verdict"] in ("clean", "unknown"):
        print("  Lưu ý: chưa bị gắn cờ KHÔNG có nghĩa là an toàn — domain mới/ít traffic thường")
        print("  chưa kịp bị cộng đồng báo cáo. Vẫn cần xác minh nội dung trang thủ công.")

    print("\n--- Khuyến nghị kênh báo cáo ---")
    print("  1. Google Safe Browsing (submit thủ công, chưa có API report):")
    print("     https://safebrowsing.google.com/safebrowsing/report_phish/")
    print("     Nội dung mô tả mẫu (copy vào ô mô tả của form):")
    print(f"     \"{generate_safebrowsing_report_text(domain, cfg)}\"")
    print("     Microsoft SmartScreen: report thủ công tại")
    print("     https://www.microsoft.com/wdsi/support/report-unsafe-site-guest/")

    if cf or cdn_detected:
        print("  2. CDN:")
        if cf:
            print(f"     - Cloudflare (form thủ công): {CDN_ABUSE_CONTACTS['cloudflare']['report_url']}")
            print(f"       {CDN_ABUSE_CONTACTS['cloudflare']['note']}")
            print("       Nội dung mô tả mẫu (copy vào ô mô tả của form):")
            print(f"       \"{generate_cloudflare_report_text(domain, cfg)}\"")
        for name in cdn_detected:
            info = CDN_ABUSE_CONTACTS.get(name)
            if info:
                print(f"     - {name.title()}: {info['report_url']}")

    print("  3. Cộng đồng bảo mật:")
    if result["virustotal_submit"]:
        print(f"     - VirusTotal: {result['virustotal_submit']}")
    elif vt.get("link"):
        print(f"     - VirusTotal: {vt['link']}")
    else:
        print("     - VirusTotal: chưa từng quét domain này — thêm cờ --submit để tự động submit")
    print("     - PhishTank: report thủ công tại https://phishtank.org/")
    print(f"     - OpenPhish: không còn hoạt động. Thay bằng APWG — draft tại reports/{domain}_apwg_report.txt")

    if ca_note:
        print(f"  4. CA ({ca_note['ca']}): {ca_note['note']}")
        if ca_note["report_url"]:
            print(f"     Report tại: {ca_note['report_url']}")
    if who.get("registrar"):
        print(f"  5. Registrar: {who['registrar']} — email: {who.get('emails')}")

    registry_contact = result["registry_contact"]
    if registry_contact.get("source") != "not_found":
        print("  6. Registry (ccTLD) — leo thang khi registrar không phản hồi:")
        if registry_contact["source"] == "static_table":
            print(f"     {registry_contact.get('registry')}: {registry_contact.get('abuse_email')}")
            if registry_contact.get("note"):
                print(f"     Lưu ý: {registry_contact['note']}")
        else:
            print(f"     WHOIS server: {registry_contact.get('whois_server')} (qua IANA referral)")
            print("     — xem draft để đọc abuse email từ nội dung WHOIS thô")
        print(f"     Draft đã tạo sẵn tại reports/{domain}_registry_report.txt")
    else:
        print(f"  6. Registry (ccTLD): TLD .{domain.rsplit('.', 1)[-1]} chưa có trong danh sách hỗ trợ,")
        print("     tra thủ công tại https://www.iana.org/domains/root/db")

    print(f"  7. VNCERT: draft đã tạo sẵn tại reports/{domain}_vncert_report.txt")
    print("     (CHỈ gửi nếu domain nhắm vào nạn nhân tại Việt Nam — tool không tự xác định được điều này)")

    origin_ip_whois = result["origin_ip_whois"]
    if origin_ip_whois:
        ip, ipw = next(iter(origin_ip_whois.items()))
        print(f"  8. Hosting/ISP (IP gốc {ip}):")
        if "error" in ipw:
            print(f"     Lỗi tra IP WHOIS: {ipw['error']}")
        else:
            print(f"     Tổ chức: {ipw.get('org') or 'N/A'}")
            print(f"     Abuse email: {ipw.get('abuse_email') or 'N/A'}")
            print(f"     ASN: {ipw.get('asn')} ({ipw.get('asn_description')})")
        print(f"     Draft đã tạo sẵn tại reports/{domain}_hosting_report.txt")
    else:
        print("  8. Hosting/ISP (IP gốc): Không phát hiện IP gốc khác qua subdomain scan — có thể")
        print("     domain không dùng CDN, hoặc IP gốc không lộ qua các subdomain thông dụng đã quét.")

    origin_candidates = {
        sub: ip for sub, ip in origin_ip_scan.items() if ip and ip != cert.get("ip")
    }
    if origin_candidates:
        print("\n--- Khả năng lộ IP gốc (candidate origin IP, cần xác minh thêm) ---")
        for sub, ip in origin_candidates.items():
            print(f"  - {sub}.{domain}: {ip}")
        print("  Đây chỉ là gợi ý dựa trên IP khác IP chính của domain, KHÔNG phải kết luận chắc")
        print("  chắn — cần xác minh thêm trước khi dùng để report hosting.")

    if result["log_error"]:
        print(f"\nLỖI khi ghi log vào {LOG_PATH}: {result['log_error']}")
    else:
        print(f"\nĐã ghi log vào {LOG_PATH}")

    if result["drafts_error"]:
        print(f"\nLỖI khi sinh email báo cáo: {result['drafts_error']}")
    elif result["drafts"]:
        print("\nĐã tạo sẵn email báo cáo:")
        for d in result["drafts"]:
            print(f"  - {d}")


def cmd_related(args):
    print(f"\n=== Tìm domain liên quan tới từ khóa: {args.keyword} ===\n")
    res = crtsh_related(args.keyword)
    if "error" in res:
        print(f"Lỗi: {res['error']}")
        return
    print(f"Tìm thấy {res['count']} domain có chứng chỉ chứa '{args.keyword}':\n")
    for d in res["domains"]:
        print(f"  - {d}")


def cmd_brandscan(args):
    domain = normalize_domain(args.target)
    print(f"\n=== Quét biến thể domain giống: {domain} (có thể mất vài phút, brand phổ biến có thể tới 10 phút) ===\n")
    res = brand_scan(domain, args.limit)
    if "error" in res:
        print(f"Lỗi: {res['error']}")
        return
    print(f"Tìm thấy {res['count']} domain biến thể ĐÃ ĐĂNG KÝ:\n")
    for r in res["results"]:
        print(f"  - {r.get('domain')}  (fuzzer: {r.get('fuzzer')})  dns_a: {r.get('dns_a')}")


def cmd_send(args):
    """CLI gửi email đồng loạt qua tất cả smtp_accounts — gửi ngay không hỏi xác nhận."""
    cfg = load_config()
    path = args.draft_path
    if not os.path.isfile(path):
        print(f"Không tìm thấy file: {path}")
        return

    parsed = parse_draft_email(path)
    if not parsed["to"]:
        print("Draft này không có địa chỉ email hợp lệ để gửi tự động (có thể cần report qua")
        print("form web, hoặc chưa tra được abuse email) — xem nội dung file và report thủ công.")
        return

    accounts = cfg.get("smtp_accounts", [])
    proxies = cfg.get("smtp_proxies", [])
    n_acc = len(accounts)
    n_proxy = len(proxies)
    print(f"\n--- Gửi email đồng loạt ---")
    print(f"To      : {parsed['to']}")
    print(f"Subject : {parsed['subject']}")
    print(f"Tài khoản: {n_acc}  |  Proxy: {n_proxy}")
    if proxies:
        for i, acc in enumerate(accounts):
            print(f"  [{i+1}] {acc.get('username')}  →  proxy {proxies[i % len(proxies)]}")
    else:
        for i, acc in enumerate(accounts):
            print(f"  [{i+1}] {acc.get('username')}  (không proxy)")

    results = send_report_email_bulk(parsed["to"], parsed["subject"], parsed["body"], cfg)
    ts = datetime.now(timezone.utc).isoformat()
    for r in results:
        status = "✓ OK" if r["success"] else f"✗ LỖI: {r['error']}"
        print(f"  {r['account']}  [{r['proxy']}]  →  {status}")
        try:
            log_sent({
                "timestamp": ts,
                "domain": domain_from_draft_filename(os.path.basename(path)),
                "draft_file": os.path.basename(path),
                "to": parsed["to"],
                "subject": parsed["subject"],
                "account": r.get("account") or "",
                "success": r["success"],
                "error": r.get("error") or "",
            })
        except Exception as e:
            print(f"  Cảnh báo: ghi {SENT_LOG_PATH} lỗi: {e}")


def main():
    parser = argparse.ArgumentParser(description="Bộ công cụ phát hiện & báo cáo phishing")
    sub = parser.add_subparsers(dest="command", required=True)

    p_check = sub.add_parser("check", help="Kiểm tra đầy đủ 1 domain")
    p_check.add_argument("target", help="Domain hoặc URL")
    p_check.add_argument("--submit", action="store_true", help="Submit lên VirusTotal nếu chưa có dữ liệu")
    p_check.set_defaults(func=cmd_check)

    p_related = sub.add_parser("related", help="Tìm domain liên quan qua crt.sh")
    p_related.add_argument("keyword", help="Từ khóa / tên thương hiệu")
    p_related.set_defaults(func=cmd_related)

    p_brand = sub.add_parser("brandscan", help="Quét biến thể domain (dnstwist)")
    p_brand.add_argument("target", help="Domain gốc hợp pháp cần bảo vệ")
    p_brand.add_argument("--limit", type=int, default=50)
    p_brand.set_defaults(func=cmd_brandscan)

    p_send = sub.add_parser("send", help="Gửi 1 draft email đã sinh sẵn qua SMTP thật (cần xác nhận thủ công)")
    p_send.add_argument("draft_path", help="Đường dẫn tới file draft trong reports/")
    p_send.set_defaults(func=cmd_send)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
